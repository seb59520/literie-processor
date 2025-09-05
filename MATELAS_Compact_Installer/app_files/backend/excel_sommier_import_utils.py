#!/usr/bin/env python3
"""
Script d'automatisation Excel pour l'import de configurations de sommiers
"""

import os
import openpyxl
import logging
from typing import List, Dict, Optional, Tuple
from datetime import datetime

# Configuration du logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelSommierImporter:
    """
    Classe pour automatiser l'import de configurations de sommiers dans Excel
    """
    
    def __init__(self, template_path: str = "template/template_sommier.xlsx", 
                 alignment_mode: str = "intelligent"):
        """
        Initialise l'importateur de sommiers
        
        Args:
            template_path: Chemin vers le template Excel sommier
            alignment_mode: Mode d'alignement ("intelligent" ou "standard")
        """
        self.template_path = template_path
        self.alignment_mode = alignment_mode
        
        # Charger le mapping manager pour utiliser les mappings sauvegard√©s
        try:
            from mapping_manager import MappingManager
            self.mapping_manager = MappingManager()
            logger.info("Mapping manager charg√© avec succ√®s pour les sommiers")
        except ImportError as e:
            logger.warning(f"Impossible de charger le mapping manager: {e}. Utilisation des mappings par d√©faut.")
            self.mapping_manager = None
        
        # Configuration pour les sommiers (m√™me logique que les matelas)
        self.column_blocks = [
            ('C', 'D'),  # Cas 1
            ('E', 'F'),  # Cas 2
            ('G', 'H'),  # Cas 3
            ('I', 'J'),  # Cas 4
            ('K', 'L'),  # Cas 5
            # M & N sont verrouill√©s - on saute
            ('O', 'P'),  # Cas 6
            ('Q', 'R'),  # Cas 7
            ('S', 'T'),  # Cas 8
            ('U', 'V'),  # Cas 9
            ('W', 'X'),  # Cas 10
        ]
        
        # Mapping par d√©faut pour les sommiers (utilis√© si mapping_manager non disponible)
        self.default_sommier_mappings = {
            # Informations client (m√™mes que les matelas)
            "Client_D1": "E1",
            "Adresse_D3": "E3",
            "numero_D2": "E2",
            
            # Champs commande et dates
            "semaine_D5": "E5",
            "lundi_D6": "E6",
            "vendredi_D7": "E7",
            
            # Champs sommiers
            "Type_Sommier_D20": "E20",
            "Materiau_D25": "E25",
            "Hauteur_D30": "E30",
            "Dimensions_D35": "E35",
            "Quantite_D40": "E40",
            # Nouvelles caract√©ristiques
            "Sommier_DansUnLit_D45": "E45",
            "Sommier_Pieds_D50": "E50",
        }
        
        self.max_cases_per_file = 10  # Nombre maximum de sommiers par fichier (m√™me que les matelas)
        
        # √âtat du fichier actuel
        self.current_workbook = None
        self.current_worksheet = None
        self.current_file_index = 1
        self.current_case_count = 0
    
    def generate_filename(self, semaine: str, id_client: str) -> str:
        """
        G√©n√®re le nom de fichier pour les sommiers
        
        Args:
            semaine: Code semaine (ex: S01)
            id_client: ID du client
            
        Returns:
            Nom de fichier g√©n√©r√©
        """
        return f"Sommier_{semaine}_{id_client}_{self.current_file_index}.xlsx"
    
    def load_template(self) -> openpyxl.Workbook:
        """
        Charge le template sommier
        
        Returns:
            Workbook du template
        """
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template sommier introuvable: {self.template_path}")
        
        logger.info(f"Chargement du template sommier: {self.template_path}")
        return openpyxl.load_workbook(self.template_path)
    
    def is_block_empty(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                      left_col: str, right_col: str) -> bool:
        """
        V√©rifie si un bloc de colonnes est vide pour les sommiers
        
        Args:
            worksheet: Feuille de calcul
            left_col: Colonne de gauche
            right_col: Colonne de droite
            
        Returns:
            True si le bloc est vide
        """
        # V√©rifier les cellules cl√©s pour les sommiers
        # On ignore les num√©ros de cas du template (1, 2, 3, etc.)
        key_cells = [
            f"{left_col}1",  # Client
            f"{left_col}3",  # Adresse
            f"{left_col}5",  # Semaine
            f"{left_col}20", # Type sommier
        ]
        
        for cell_address in key_cells:
            cell_value = worksheet[cell_address].value
            if cell_value and str(cell_value).strip():
                # V√©rifier que ce n'est pas juste un num√©ro de cas du template
                cell_str = str(cell_value).strip()
                if not cell_str.isdigit() or int(cell_str) > 20:  # Ignorer les num√©ros 1-20 du template
                    return False
        
        return True
    
    def find_next_empty_block(self, worksheet: openpyxl.worksheet.worksheet.Worksheet) -> Optional[Tuple[str, str]]:
        """
        Trouve le prochain bloc vide pour les sommiers
        
        Args:
            worksheet: Feuille de calcul
            
        Returns:
            Tuple (colonne_gauche, colonne_droite) ou None
        """
        for left_col, right_col in self.column_blocks:
            if self.is_block_empty(worksheet, left_col, right_col):
                logger.info(f"V√©rification bloc {left_col}-{right_col}: cellule {left_col}1 est vide ou fusionn√©e")
                logger.info(f"Bloc vide trouv√©: {left_col}-{right_col}")
                return (left_col, right_col)
        return None
    
    def map_json_to_cells(self, config_json: Dict, left_col: str, right_col: str) -> Dict[str, str]:
        """
        Mappe les donn√©es JSON vers les cellules Excel pour les sommiers
        
        Args:
            config_json: Configuration JSON du sommier
            left_col: Colonne de gauche
            right_col: Colonne de droite
            
        Returns:
            Dictionnaire {adresse_cellule: valeur}
        """
        cell_mapping = {}
        
        # Utiliser le mapping manager si disponible, sinon le mapping par d√©faut
        if self.mapping_manager:
            # R√©cup√©rer les mappings sauvegard√©s pour les sommiers
            saved_mappings = self.mapping_manager.sommiers_mappings
            logger.info(f"Utilisation des mappings sauvegard√©s pour les sommiers: {len(saved_mappings)} champs")
        else:
            saved_mappings = self.default_sommier_mappings
            logger.info("Utilisation des mappings par d√©faut pour les sommiers")

        # Convertir les mappings sauvegard√©s au format attendu
        for field_name, cell_address in saved_mappings.items():
            if not cell_address:  # Champ ignor√©
                continue
                
            # Extraire colonne et ligne de l'adresse de cellule
            import re
            match = re.match(r'([A-Z]+)(\d+)', cell_address)
            if match:
                col_letter = match.group(1)
                row_num = int(match.group(2))
                
                # D√©terminer si c'est une colonne C ou D dans le template
                if col_letter in ['C', 'E', 'G', 'I', 'K', 'O', 'Q', 'S', 'U', 'W']:
                    actual_col = left_col
                else:
                    actual_col = right_col
                
                actual_cell_address = f"{actual_col}{row_num}"
                value = config_json.get(field_name, "")
                
                # Ne pas √©crire les champs vides (ignor√©s)
                if value:
                    cell_mapping[actual_cell_address] = str(value)
                    logger.debug(f"Mapping sommier: {field_name} -> {actual_cell_address} = {value}")
        
        return cell_mapping
    
    def apply_cell_alignment(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                           cell_address: str, json_key: str) -> None:
        """
        Applique l'alignement intelligent aux cellules pour les sommiers
        
        Args:
            worksheet: Feuille de calcul
            cell_address: Adresse de la cellule
            json_key: Cl√© JSON correspondante
        """
        if self.alignment_mode == "intelligent":
            # Alignement sp√©cifique aux sommiers
            if "Client" in json_key:
                worksheet[cell_address].alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")
            elif "Article" in json_key:
                worksheet[cell_address].alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")
            elif "Quantite" in json_key:
                worksheet[cell_address].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            elif "Dimensions" in json_key:
                worksheet[cell_address].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            elif "Prix" in json_key:
                worksheet[cell_address].alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")
            else:
                worksheet[cell_address].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    
    def center_block_cells(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                          left_col: str, right_col: str) -> None:
        """
        Centre toutes les cellules d'un bloc pour les sommiers
        
        Args:
            worksheet: Feuille de calcul
            left_col: Colonne de gauche
            right_col: Colonne de droite
        """
        for row in range(1, 50):  # Lignes 1 √† 50
            for col in range(ord(left_col), ord(right_col) + 1):
                cell_address = f"{chr(col)}{row}"
                try:
                    worksheet[cell_address].alignment = openpyxl.styles.Alignment(
                        horizontal="center", 
                        vertical="center"
                    )
                except:
                    pass  # Ignorer les erreurs de cellules non existantes
    
    def update_case_numbers(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                           file_index: int) -> None:
        """
        Met √† jour les num√©ros de cas dans le fichier sommier selon l'index du fichier
        
        Args:
            worksheet: Feuille de calcul
            file_index: Index du fichier (1, 2, 3, etc.)
        """
        try:
            # Calculer le num√©ro de d√©part pour ce fichier
            # Fichier 1: cas 1-10, Fichier 2: cas 11-20, Fichier 3: cas 21-30, etc.
            start_case_number = (file_index - 1) * self.max_cases_per_file + 1
            
            logger.info(f"Mise √† jour des num√©ros de cas sommier pour le fichier {file_index}: cas {start_case_number} √† {start_case_number + 9}")
            
            # Mapping des colonnes pour les num√©ros de cas (ligne 2)
            # Format: (colonne_gauche, colonne_droite, num√©ro_cas)
            case_number_mapping = [
                ('C', 'D', start_case_number),      # Cas 1
                ('E', 'F', start_case_number + 1),  # Cas 2
                ('G', 'H', start_case_number + 2),  # Cas 3
                ('I', 'J', start_case_number + 3),  # Cas 4
                ('K', 'L', start_case_number + 4),  # Cas 5
                # M & N sont verrouill√©s - on saute
                ('O', 'P', start_case_number + 5),  # Cas 6
                ('Q', 'R', start_case_number + 6),  # Cas 7
                ('S', 'T', start_case_number + 7),  # Cas 8
                ('U', 'V', start_case_number + 8),  # Cas 9
                ('W', 'X', start_case_number + 9),  # Cas 10
            ]
            
            # Mettre √† jour chaque num√©ro de cas
            for left_col, right_col, case_number in case_number_mapping:
                # Mettre √† jour la colonne gauche (C, E, G, etc.)
                left_cell = f"{left_col}2"
                worksheet[left_cell] = case_number
                
                # Mettre √† jour la colonne droite (D, F, H, etc.)
                right_cell = f"{right_col}2"
                worksheet[right_cell] = case_number
                
                logger.debug(f"Cas sommier {case_number}: {left_cell} et {right_cell} mis √† jour")
            
            logger.info(f"Num√©ros de cas sommier mis √† jour avec succ√®s pour le fichier {file_index}")
            
        except Exception as e:
            logger.error(f"Erreur lors de la mise √† jour des num√©ros de cas sommier: {e}")
            raise

    def write_config_to_block(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                             config_json: Dict, left_col: str, right_col: str) -> None:
        """
        √âcrit une configuration sommier dans un bloc sp√©cifique
        
        Args:
            worksheet: Feuille de calcul
            config_json: Configuration JSON du sommier
            left_col: Colonne de gauche
            right_col: Colonne de droite
        """
        # Mapper les donn√©es vers les cellules
        cell_mappings = self.map_json_to_cells(config_json, left_col, right_col)
        
        # √âcrire les donn√©es
        for cell_address, value in cell_mappings.items():
            worksheet[cell_address] = value
            logger.info(f"√âcriture: {cell_address} = {value}")
            
            # Appliquer l'alignement
            json_key = next((k for k, v in cell_mappings.items() if v == value), "")
            self.apply_cell_alignment(worksheet, cell_address, json_key)
        
        # Centrer toutes les cellules du bloc
        self.center_block_cells(worksheet, left_col, right_col)

    def create_new_file(self, semaine: str, id_fichier: str) -> openpyxl.Workbook:
        """
        Cr√©e un nouveau fichier Excel sommier √† partir du template
        
        Args:
            semaine: Code semaine
            id_fichier: ID du fichier
            
        Returns:
            Nouveau workbook
        """
        self.current_file_index += 1
        self.current_case_count = 0
        
        logger.info(f"Cr√©ation nouveau fichier sommier: index {self.current_file_index}")
        print(f"DEBUG SOMMIER: create_new_file appel√©e avec semaine={semaine}")
        
        # Charge le template pour le nouveau fichier
        new_workbook = self.load_template()
        new_worksheet = new_workbook.active
        
        # Met √† jour les num√©ros de cas pour ce nouveau fichier
        self.update_case_numbers(new_worksheet, self.current_file_index)
        
        # Ajouter le num√©ro de semaine en A4
        try:
            # Extraire le num√©ro de semaine (enlever le 'S' du d√©but)
            numero_semaine = semaine.replace('S', '') if semaine.startswith('S') else semaine
            semaine_text = f"sem. {numero_semaine}"
            
            # Forcer l'√©criture en A4, m√™me si cellules fusionn√©es
            new_worksheet['A4'].value = semaine_text
            
            # Si A4 et B4 sont fusionn√©es, s'assurer que B4 est vide
            if 'B4' in new_worksheet:
                new_worksheet['B4'].value = None
            
            logger.info(f"Num√©ro de semaine forc√© en A4 (sommier): {semaine_text}")
            print(f"DEBUG SOMMIER: A4 √©crit avec succ√®s: {semaine_text}")
        except Exception as e:
            logger.warning(f"Erreur lors de l'ajout du num√©ro de semaine en A4 (sommier): {e}")
            print(f"DEBUG SOMMIER: Erreur A4 principale: {e}")
            # Tentative alternative : √©criture directe sans v√©rification
            try:
                new_worksheet.cell(row=4, column=1, value=f"sem. {numero_semaine}")
                logger.info(f"Num√©ro de semaine √©crit via cell() en A4 (sommier): sem. {numero_semaine}")
                print(f"DEBUG SOMMIER: A4 √©crit via fallback: sem. {numero_semaine}")
            except Exception as e2:
                logger.error(f"√âchec total de l'√©criture en A4 (sommier): {e2}")
                print(f"DEBUG SOMMIER: √âchec total A4: {e2}")
        
        return new_workbook
    
    def save_workbook(self, workbook: openpyxl.Workbook, semaine: str, id_fichier: str) -> str:
        """
        Sauvegarde le workbook sommier avec le nom correct
        
        Args:
            workbook: Workbook √† sauvegarder
            semaine: Code semaine
            id_fichier: ID du fichier
            
        Returns:
            Chemin du fichier sauvegard√©
        """
        filename = self.generate_filename(semaine, id_fichier)
        
        # Utiliser le r√©pertoire de sortie configur√©
        try:
            from config import config
            output_dir = config.get_excel_output_directory()
        except Exception as e:
            logger.warning(f"Impossible d'importer config: {e}. Utilisation du r√©pertoire par d√©faut.")
            output_dir = "output"
        
        filepath = os.path.join(output_dir, filename)
        
        # Cr√©e le dossier de sortie s'il n'existe pas
        os.makedirs(output_dir, exist_ok=True)
        
        # Auto-ajuster la largeur des colonnes C, E, G, I, K, M, O, Q, S, U, W bas√© sur le contenu de la ligne 2
        if workbook.active:
            ws = workbook.active
            columns_to_adjust = ['C', 'E', 'G', 'I', 'K', 'M', 'O', 'Q', 'S', 'U', 'W']
            
            for col_letter in columns_to_adjust:
                try:
                    # R√©cup√©rer le contenu de la cellule en ligne 2 (ex: C2, E2, G2...)
                    cell_value = ws[f'{col_letter}2'].value
                    if cell_value:
                        # Calculer la largeur bas√©e sur le contenu exact + petite marge
                        content_length = len(str(cell_value))
                        # Largeur minimale 3, avec juste 1 caract√®re de marge
                        adjusted_width = max(content_length + 1, 3)
                        ws.column_dimensions[col_letter].width = adjusted_width
                        logger.info(f"üìè Sommier colonne {col_letter}: largeur ajust√©e √† {adjusted_width} (contenu: '{str(cell_value)}' = {content_length} chars)")
                except Exception as e:
                    logger.warning(f"‚ö†Ô∏è Erreur ajustement sommier colonne {col_letter}: {e}")
        
        workbook.save(filepath)
        logger.info(f"Fichier sommier sauvegard√©: {filepath}")
        
        return filepath
    
    def import_configurations(self, configurations: List[Dict], semaine: str, id_fichier: str) -> List[str]:
        """
        Importe une liste de configurations sommier dans Excel
        
        Args:
            configurations: Liste des configurations JSON
            semaine: Code semaine
            id_fichier: ID du fichier
            
        Returns:
            Liste des fichiers cr√©√©s
        """
        created_files = []

        # Utiliser le r√©pertoire de sortie configur√©
        try:
            from config import config
            output_dir = config.get_excel_output_directory()
        except Exception as e:
            logger.warning(f"Impossible d'importer config: {e}. Utilisation du r√©pertoire par d√©faut.")
            output_dir = "output"
        
        # Cherche le dernier fichier existant pour cette semaine
        self.current_file_index = 1
        while True:
            filename = self.generate_filename(semaine, id_fichier)
            filepath = os.path.join(output_dir, filename)
            if os.path.exists(filepath):
                self.current_file_index += 1
            else:
                break
        
        # Si on a trouv√© des fichiers existants, charge le dernier
        if self.current_file_index > 1:
            self.current_file_index -= 1
            filename = self.generate_filename(semaine, id_fichier)
            filepath = os.path.join(output_dir, filename)
            logger.info(f"Chargement du fichier sommier existant: {filepath}")
            self.current_workbook = openpyxl.load_workbook(filepath)
            self.current_worksheet = self.current_workbook.active
            
            # Compte les cas d√©j√† pr√©sents
            self.current_case_count = 0
            for left_col, right_col in self.column_blocks:
                if not self.is_block_empty(self.current_worksheet, left_col, right_col):
                    self.current_case_count += 1
            logger.info(f"Fichier sommier existant contient {self.current_case_count} cas")
        else:
            # Charge le template initial
            filename = self.generate_filename(semaine, id_fichier)
            filepath = os.path.join(output_dir, filename)
            logger.info(f"Cr√©ation d'un nouveau fichier sommier: {filepath}")
            self.current_workbook = self.load_template()
            self.current_worksheet = self.current_workbook.active
            self.current_case_count = 0
            
            # Ajouter le num√©ro de semaine en A4 pour le premier fichier
            try:
                numero_semaine = semaine.replace('S', '') if semaine.startswith('S') else semaine
                semaine_text = f"sem. {numero_semaine}"
                self.current_worksheet['A4'].value = semaine_text
                if 'B4' in self.current_worksheet:
                    self.current_worksheet['B4'].value = None
                logger.info(f"Num√©ro de semaine ajout√© au premier fichier sommier en A4: {semaine_text}")
                print(f"DEBUG PREMIER FICHIER SOMMIER: A4 √©crit avec succ√®s: {semaine_text}")
            except Exception as e:
                logger.error(f"Erreur lors de l'ajout du num√©ro de semaine au premier fichier sommier: {e}")
                print(f"DEBUG ERREUR PREMIER FICHIER SOMMIER: {e}")
        
        for i, config in enumerate(configurations):
            logger.info(f"Traitement configuration sommier {i+1}/{len(configurations)}")
            
            # V√©rifie si on doit cr√©er un nouveau fichier
            if self.current_case_count >= self.max_cases_per_file:
                # Sauvegarde le fichier actuel
                filepath = self.save_workbook(self.current_workbook, semaine, id_fichier)
                created_files.append(filepath)
                
                # Cr√©e un nouveau fichier
                self.current_workbook = self.create_new_file(semaine, id_fichier)
                self.current_worksheet = self.current_workbook.active
            
            # Trouve le prochain bloc vide
            empty_block = self.find_next_empty_block(self.current_worksheet)
            
            if empty_block is None:
                # Tous les blocs sont pleins, cr√©e un nouveau fichier
                filepath = self.save_workbook(self.current_workbook, semaine, id_fichier)
                created_files.append(filepath)
                
                self.current_workbook = self.create_new_file(semaine, id_fichier)
                self.current_worksheet = self.current_workbook.active
                
                # Retrouve le bloc vide dans le nouveau fichier
                empty_block = self.find_next_empty_block(self.current_worksheet)
            
            left_col, right_col = empty_block
            
            # √âcrit la configuration dans le bloc
            self.write_config_to_block(self.current_worksheet, config, left_col, right_col)
            
            self.current_case_count += 1
            logger.info(f"Configuration sommier {i+1} √©crite dans le bloc {left_col}-{right_col}")
        
        # Sauvegarde le dernier fichier
        if self.current_workbook:
            filepath = self.save_workbook(self.current_workbook, semaine, id_fichier)
            created_files.append(filepath)
        
        logger.info(f"Import sommier termin√©. {len(created_files)} fichier(s) cr√©√©(s)")
        return created_files


def main():
    """
    Fonction principale pour tester l'importateur de sommiers
    """
    # Exemple d'utilisation
    importer = ExcelSommierImporter()
    
    # Exemple de configurations JSON pour sommiers
    sample_configs = [
        {
            "Client_D1": "Client Test 1",
            "Client_D2": "Adresse Test 1",
            "Article_D6": "Sommier √† lattes",
            "Quantite_D11": "2",
            "Dimensions_D15": "160x200",
            "Type_Sommier_D20": "SOMMIER √Ä LATTES",
            "Materiau_D25": "BOIS",
            "Hauteur_D30": "8",
            "Prix_D35": "400.00"
        },
        {
            "Client_D1": "Client Test 2",
            "Client_D2": "Adresse Test 2",
            "Article_D6": "Sommier tapissier",
            "Quantite_D11": "1",
            "Dimensions_D15": "140x190",
            "Type_Sommier_D20": "SOMMIER TAPISSIER",
            "Materiau_D25": "TAPISSIER",
            "Hauteur_D30": "12",
            "Prix_D35": "300.00"
        }
    ]
    
    try:
        created_files = importer.import_configurations(sample_configs, "S01", "1234")
        print(f"Fichiers sommiers cr√©√©s: {created_files}")
    except Exception as e:
        logger.error(f"Erreur lors de l'import sommier: {e}")


if __name__ == "__main__":
    main() 