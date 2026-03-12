#!/usr/bin/env python3
"""
Script d'automatisation Excel pour l'import de configurations de sommiers
"""

import os
import re
import math
import unicodedata
import openpyxl
import logging
from typing import List, Dict, Optional, Tuple
from datetime import datetime
from openpyxl.utils import get_column_letter

# Patterns pour détecter et supprimer le boilerplate PDF de page 2
# (en-tête de Literie Westelynck qui se retrouve en fin d'article quand le PDF a 2 pages)
_PDF_BOILERPLATE_PATTERNS = [
    re.compile(r'SAS Literie Westelynck.*$', re.DOTALL | re.IGNORECASE),
    re.compile(r'Siret\s+429\s+352\s+891.*$', re.DOTALL | re.IGNORECASE),
    re.compile(r'CEE\s+FR50\s+429\s+352.*$', re.DOTALL | re.IGNORECASE),
    re.compile(r'Domiciliation\s+Bancaire.*$', re.DOTALL | re.IGNORECASE),
    re.compile(r'Montant\s+TTC\s+P\.?U\.?\s+TTC\s+Qt[eé]\s+Description.*$', re.DOTALL | re.IGNORECASE),
    re.compile(r'Maison\s*\n\s*Fond[ée]e\s+en\s*\n.*$', re.DOTALL | re.IGNORECASE),
]


def _strip_pdf_boilerplate(text: str) -> str:
    """Supprime le texte boilerplate du PDF (en-tête page 2) des descriptions d'articles."""
    if not text:
        return text
    result = text
    for pattern in _PDF_BOILERPLATE_PATTERNS:
        result = pattern.sub('', result)
    return result.strip()


FIELD_ROW_OVERRIDES = {
    # Zones lignes 6-13
    "Sommier_DansUnLit_B6": {"row": 6, "column": "left"},
    "Dimension_Sommier_C6": {"row": 6, "column": "right"},
    "Sommier_Pieds_B7": {"row": 7, "column": "left"},
    "Dimension_Sommier_C7": {"row": 7, "column": "right"},
    "Tete_Bois_C9": {"row": 9, "column": "right"},
    "Dosseret_Tissu_C10": {"row": 10, "column": "right"},
    "Jumeaux_B11": {"row": 11, "column": "left"},
    "Chevets_C13": {"row": 13, "column": "right"},
    # Finitions structure
    "Finition_Multiplis_D76": {"row": 25, "column": "left"},
    "Finition_90mm_D77": {"row": 25, "column": "right"},
    "Parementee_B26": {"row": 26, "column": "left"},
    "Parementee_Info_C26": {"row": 26, "column": "right"},
    "Finition_Paremente_D78": {"row": 26, "column": "left"},
    "Finition_Multiplis_TV_D79": {"row": 27, "column": "left"},
    "Finition_Multiplis_L_D80": {"row": 27, "column": "right"},
    "Finition_Frene_TV_D82": {"row": 28, "column": "left"},
    "Finition_Frene_L_D83": {"row": 28, "column": "right"},
    "Finition_Chene_D81": {"row": 29, "column": "left"},
    # Options pieds/patins lignes 33-43
    "Platine_Reunion_D71": {"row": 39, "column": "left"},
    "Pieds_Centraux_D72": {"row": 40, "column": "left"},
    "Patins_Feutre_D73": {"row": 41, "column": "left"},
    "Patins_Carrelage_D74": {"row": 42, "column": "left"},
    "Patins_Teflon_D75": {"row": 43, "column": "left"},
    # Options diverses lignes 45-55
    "Butees_Laterales_D60": {"row": 45, "column": "left"},
    "Butees_Pieds_D61": {"row": 46, "column": "left"},
    "Solidarisation_D62": {"row": 47, "column": "left"},
    "Demi_Corbeille_D63": {"row": 48, "column": "left"},
    "Profile_D64": {"row": 49, "column": "left"},
    "Renforces_D65": {"row": 50, "column": "left"},
    "Genou_Moins_D66": {"row": 51, "column": "left"},
    "Tronc_Plus_D67": {"row": 52, "column": "left"},
    "Calles_D68": {"row": 53, "column": "left"},
    "Rampes_D69": {"row": 54, "column": "left"},
    "Autre_D70": {"row": 55, "column": "left"},
    # Opérations logistiques lignes 56-58
    "emporte_client_C57": {"row": 56, "column": "left"},
    "fourgon_C58": {"row": 57, "column": "left"},
    "transporteur_C59": {"row": 58, "column": "left"},
}

# Configuration du logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelSommierImporter:
    """
    Classe pour automatiser l'import de configurations de sommiers dans Excel
    """

    @staticmethod
    def extraire_nom_famille(nom_complet: str) -> str:
        """
        Extrait le nom de famille depuis un nom complet.
        Ex: "Mr et Me ADISSON EMMANUEL & VERONIQUE" → "ADISSON"
        Ex: "Mr et Mme BOUDRY REGIS & CHRISTINE" → "BOUDRY"
        Ex: "Me DEVYNCK ARLETTE" → "DEVYNCK"
        """
        if not nom_complet:
            return ""

        nom = nom_complet.strip()

        # Supprimer les civilités en début
        prefixes_to_remove = [
            "Mr et Mme ", "Mr et Me ", "Mr et me ", "Mr & Mme ", "Mr & Me ",
            "M. et Mme ", "M. et Me ", "Mme et Mr ", "Me et Mr ",
            "Mme ", "Me ", "Mr ", "M. ", "Mlle ",
        ]
        for prefix in prefixes_to_remove:
            if nom.startswith(prefix):
                nom = nom[len(prefix):].strip()
                break

        # Le premier mot restant est le nom de famille
        parts = nom.split()
        if parts:
            return parts[0].upper()

        return nom.upper()

    def __init__(self, template_path: str = "template/template_sommier.xlsx",
                 alignment_mode: str = "intelligent"):
        """
        Initialise l'importateur de sommiers
        
        Args:
            template_path: Chemin vers le template Excel sommier
            alignment_mode: Mode d'alignement ("intelligent" ou "standard")
        """
        self.template_path = template_path
        self.alignment_mode = alignment_mode
        
        # Charger le mapping manager pour utiliser les mappings sauvegardés
        try:
            from mapping_manager import MappingManager
            self.mapping_manager = MappingManager()
            logger.info("Mapping manager chargé avec succès pour les sommiers")
        except ImportError as e:
            logger.warning(f"Impossible de charger le mapping manager: {e}. Utilisation des mappings par défaut.")
            self.mapping_manager = None
        
        # Configuration pour les sommiers détectée dynamiquement
        self.column_blocks = self._discover_column_blocks()
        self.field_row_overrides = FIELD_ROW_OVERRIDES.copy()
        
        # Mapping par défaut pour les sommiers (utilisé si mapping_manager non disponible)
        # Note: Les mappings sont relatifs au template, la logique de map_json_to_cells
        # applique les transformations (D→C, C→B, etc.)
        self.default_sommier_mappings = {
            # Informations client
            "Client_D1": "C1",  # Écrit en C1, E1, G1... (colonne gauche)
            "Adresse_D3": "C2",  # Écrit en C2, E2, G2... (colonne gauche)
            "numero_D2": "C4",  # Écrit en C4, E4, G4... (colonne gauche, ligne 4)
            
            # Champs commande et dates
            "semaine_D5": "A4",  # Écrit en A4 avec "sem. " (géré séparément)
        }
        
        self.max_cases_per_file = 10  # Nombre maximum de sommiers par fichier (même que les matelas)
        
        # État du fichier actuel
        self.current_workbook = None
        self.current_worksheet = None
        self.current_file_index = 1
        self.current_case_count = 0
        self.total_written_cases = 0

    def _discover_column_blocks(self) -> List[Tuple[str, str]]:
        """Analyse le template pour déterminer automatiquement les paires de colonnes B/C, D/E, etc."""
        try:
            workbook = openpyxl.load_workbook(self.template_path, data_only=True)
        except FileNotFoundError:
            logger.warning("Template introuvable pour détection auto. Utilisation du layout par défaut.")
            return [
                ('B', 'C'),
                ('D', 'E'),
                ('F', 'G'),
                ('H', 'I'),
                ('J', 'K'),
                ('M', 'N'),
                ('O', 'P'),
                ('Q', 'R'),
                ('S', 'T'),
                ('U', 'V'),
            ]

        worksheet = workbook.active
        blocks: List[Tuple[str, str]] = []
        try:
            for col_idx in range(1, worksheet.max_column):
                cell_value = worksheet.cell(row=2, column=col_idx).value
                if isinstance(cell_value, (int, float)) and str(int(cell_value)).isdigit():
                    left = get_column_letter(col_idx)
                    right = get_column_letter(col_idx + 1)
                    blocks.append((left, right))
            if not blocks:
                raise ValueError("Aucun bloc détecté dans le template sommier (ligne 2)")
            return blocks
        finally:
            workbook.close()

    def _clear_block(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, left_col: str, right_col: str) -> None:
        """Réinitialise toutes les cellules d'un bloc avant écriture."""
        for row in range(3, 90):  # préserver les en-têtes (lignes 1-2)
            worksheet[f"{left_col}{row}"] = None
            worksheet[f"{right_col}{row}"] = None
    
    def generate_filename(self, semaine: str, id_client: str) -> str:
        """
        Génère le nom de fichier pour les sommiers
        
        Args:
            semaine: Code semaine (ex: S01)
            id_client: ID du client
            
        Returns:
            Nom de fichier généré
        """
        return f"Sommier_{semaine}_{id_client}_{self.current_file_index}.xlsx"
    
    def load_template(self) -> openpyxl.Workbook:
        """
        Charge le template sommier
        
        Returns:
            Workbook du template
        """
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template sommier introuvable: {self.template_path}")
        
        logger.info(f"Chargement du template sommier: {self.template_path}")
        return openpyxl.load_workbook(self.template_path)
    
    def is_block_empty(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                      left_col: str, right_col: str) -> bool:
        """
        Vérifie si un bloc de colonnes est vide pour les sommiers
        
        Args:
            worksheet: Feuille de calcul
            left_col: Colonne de gauche
            right_col: Colonne de droite
            
        Returns:
            True si le bloc est vide
        """
        # Vérifier les cellules clés pour les sommiers
        # Les données sont écrites dans right_col (C, E, G...) pour la plupart des champs
        # On ignore les numéros de cas du template (1, 2, 3, etc.)
        key_cells = [
            f"{right_col}1",  # Client (C1, E1, G1...)
            f"{right_col}2",  # Adresse (C2, E2, G2...)
            f"{right_col}4",  # Numéro commande (C4, E4, G4...)
            f"{right_col}20", # Type sommier (si présent)
        ]
        
        for cell_address in key_cells:
            try:
                cell_value = worksheet[cell_address].value
                if cell_value and str(cell_value).strip():
                    # Vérifier que ce n'est pas juste un numéro de cas du template
                    cell_str = str(cell_value).strip()
                    # Ignorer les numéros de cas du template (1-20) et les cellules fusionnées
                    if cell_str.isdigit():
                        case_num = int(cell_str)
                        if case_num <= 200:  # Numéro de cas du template, on ignore
                            continue
                    # Si c'est autre chose qu'un numéro de cas, le bloc n'est pas vide
                    logger.info(f"Bloc {left_col}-{right_col} non vide: {cell_address} = '{cell_value}'")
                    return False
            except Exception as e:
                # Si erreur (cellule fusionnée, etc.), on considère comme vide
                logger.debug(f"Erreur vérification {cell_address}: {e}, considéré comme vide")
                continue
        
        logger.info(f"Bloc {left_col}-{right_col} est vide")
        return True
    
    def find_next_empty_block(self, worksheet: openpyxl.worksheet.worksheet.Worksheet) -> Optional[Tuple[str, str]]:
        """
        Trouve le prochain bloc vide pour les sommiers
        
        Args:
            worksheet: Feuille de calcul
            
        Returns:
            Tuple (colonne_gauche, colonne_droite) ou None
        """
        for left_col, right_col in self.column_blocks:
            if self.is_block_empty(worksheet, left_col, right_col):
                logger.info(f"Vérification bloc {left_col}-{right_col}: cellule {left_col}1 est vide ou fusionnée")
                logger.info(f"Bloc vide trouvé: {left_col}-{right_col}")
                return (left_col, right_col)
        return None
    
    def map_json_to_cells(self, config_json: Dict, left_col: str, right_col: str) -> Dict[str, str]:
        """
        Mappe les données JSON vers les cellules Excel pour les sommiers
        
        Args:
            config_json: Configuration JSON du sommier
            left_col: Colonne de gauche
            right_col: Colonne de droite
            
        Returns:
            Dictionnaire {adresse_cellule: valeur}
        """
        cell_mapping = {}
        
        # Utiliser le mapping manager si disponible, sinon le mapping par défaut
        if self.mapping_manager:
            # Récupérer les mappings sauvegardés pour les sommiers
            saved_mappings = self.mapping_manager.sommiers_mappings
            logger.info(f"Utilisation des mappings sauvegardés pour les sommiers: {len(saved_mappings)} champs")
        else:
            saved_mappings = self.default_sommier_mappings
            logger.info("Utilisation des mappings par défaut pour les sommiers")

        # Convertir les mappings sauvegardés au format attendu
        for field_name, cell_address in saved_mappings.items():
            if not cell_address:  # Champ ignoré
                continue
                
            # Extraire colonne et ligne de l'adresse de cellule
            import re
            match = re.match(r'([A-Z]+)(\d+)', cell_address)
            if match:
                col_letter = match.group(1)
                row_num = int(match.group(2))
                
                row_override = self.field_row_overrides.get(field_name)
                actual_col = None

                # Règles spéciales pour certains champs
                if row_override:
                    row_num = row_override["row"]
                    column_selector = row_override.get("column", "right")
                    if column_selector == "left":
                        actual_col = left_col
                    elif column_selector == "right":
                        actual_col = right_col
                    elif column_selector == "fixed":
                        actual_col = row_override.get("column_letter", right_col)
                    else:
                        actual_col = right_col
                elif field_name == "numero_D2":
                    # numero_D2 : toujours ligne 4, colonne droite du bloc (C4, E4, G4...)
                    actual_col = right_col
                    row_num = 4
                elif field_name == "Client_D1":
                    # Client_D1 : ligne 1, colonne droite du bloc (C1, E1, G1...)
                    actual_col = right_col
                    row_num = 1
                elif field_name == "Adresse_D3":
                    # Adresse_D3 : ligne 2, colonne droite du bloc (C2, E2, G2...)
                    actual_col = right_col
                    row_num = 2
                elif field_name == "semaine_D5":
                    # semaine_D5 : A4 avec "sem. " devant (géré ailleurs, on ignore ici)
                    continue
                else:
                    # Pour les autres champs, transformation selon les nouvelles paires BC, DE, FG...
                    if col_letter in ['C', 'E', 'G', 'I', 'K', 'O', 'Q', 'S', 'U', 'W']:
                        actual_col = left_col
                    else:
                        actual_col = right_col
                
                actual_cell_address = f"{actual_col}{row_num}"
                value = config_json.get(field_name, "")
                
                # Vérifier si PAREMENTEE est détecté pour forcer MULTIPLIS à vide
                if field_name in ['Finition_Multiplis_D76', 'Finition_Multiplis_TV_D79', 'Finition_Multiplis_L_D80']:
                    description_sommier = config_json.get("description", "") or config_json.get("article_description", "")
                    if description_sommier:
                        desc_upper = description_sommier.upper()
                        desc_normalized = description_sommier.upper().replace('É', 'E').replace('Ê', 'E')
                        if ("STRUCTURE PAREMENTEE" in desc_normalized or 
                            "STRUCTURE PAREMENTÉE" in desc_upper or 
                            "PAREMENTE" in desc_normalized or
                            "PAREMENTÉE" in desc_upper):
                            # PAREMENTEE détecté, forcer MULTIPLIS à vide
                            value = ''
                            logger.info(f"MULTIPLIS forcé à vide dans map_json_to_cells: {field_name} -> {actual_cell_address}")
                
                # Ne pas écrire les champs vides (ignorés)
                if value:
                    cell_mapping[actual_cell_address] = str(value)
                    logger.debug(f"Mapping sommier: {field_name} -> {actual_cell_address} = {value}")
        
        return cell_mapping
    
    def apply_cell_alignment(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                           cell_address: str, json_key: str) -> None:
        """
        Applique l'alignement intelligent aux cellules pour les sommiers
        
        Args:
            worksheet: Feuille de calcul
            cell_address: Adresse de la cellule
            json_key: Clé JSON correspondante
        """
        if self.alignment_mode == "intelligent":
            # Alignement spécifique aux sommiers
            if "Client" in json_key:
                worksheet[cell_address].alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")
            elif "Article" in json_key:
                worksheet[cell_address].alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")
            elif "Quantite" in json_key:
                worksheet[cell_address].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            elif "Dimensions" in json_key:
                worksheet[cell_address].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            elif "Prix" in json_key:
                worksheet[cell_address].alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")
            else:
                worksheet[cell_address].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    
    def center_block_cells(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                          left_col: str, right_col: str) -> None:
        """
        Centre toutes les cellules d'un bloc pour les sommiers
        
        Args:
            worksheet: Feuille de calcul
            left_col: Colonne de gauche
            right_col: Colonne de droite
        """
        for row in range(1, 50):  # Lignes 1 à 50
            for col in range(ord(left_col), ord(right_col) + 1):
                cell_address = f"{chr(col)}{row}"
                try:
                    worksheet[cell_address].alignment = openpyxl.styles.Alignment(
                        horizontal="center", 
                        vertical="center"
                    )
                except:
                    pass  # Ignorer les erreurs de cellules non existantes
    
    def update_case_numbers(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                           file_index: int) -> None:
        """
        Met à jour les numéros de cas dans le fichier sommier selon l'index du fichier
        
        Args:
            worksheet: Feuille de calcul
            file_index: Index du fichier (1, 2, 3, etc.)
        """
        try:
            # Calculer le numéro de départ pour ce fichier
            # Fichier 1: cas 1-10, Fichier 2: cas 11-20, Fichier 3: cas 21-30, etc.
            start_case_number = (file_index - 1) * self.max_cases_per_file + 1
            
            logger.info(f"Mise à jour des numéros de cas sommier pour le fichier {file_index}: cas {start_case_number} à {start_case_number + 9}")
            
            case_blocks = self.column_blocks[:self.max_cases_per_file]
            for offset, (left_col, right_col) in enumerate(case_blocks):
                case_number = start_case_number + offset
                left_cell = f"{left_col}2"
                right_cell = f"{right_col}2"
                worksheet[left_cell] = case_number
                worksheet[right_cell] = case_number
                logger.debug("Cas sommier %s: %s et %s mis à jour", case_number, left_cell, right_cell)
            
            logger.info(f"Numéros de cas sommier mis à jour avec succès pour le fichier {file_index}")
            
        except Exception as e:
            logger.error(f"Erreur lors de la mise à jour des numéros de cas sommier: {e}")
            raise

    def detecter_type_pieds(self, description: str) -> Optional[str]:
        """
        Détecte le type de pieds depuis une description d'article.
        
        Args:
            description: Description de l'article (ex: "JEU DE 8 PIEDS CUBIQUE BRUT 20 CM + PLATINE DE RÉUNION + PATINS TÉFLON")
            
        Returns:
            Type de pied détecté: "ANGLE_CUBIQUE", "ANGLE_DROIT", "ANGLE_GALBE", "CUBIQUE", "CYLINDRE", ou None
        """
        if not description:
            return None
        
        # Normaliser la description en majuscules
        desc_upper = description.upper()
        
        # Ordre important : chercher d'abord les types composés pour éviter les confusions
        # ANGLE CUBIQUE (ne pas confondre avec CUBIQUE seul)
        if "ANGLE CUBIQUE" in desc_upper:
            return "ANGLE_CUBIQUE"
        
        # ANGLE DROIT
        if "ANGLE DROIT" in desc_upper:
            return "ANGLE_DROIT"
        
        # ANGLE GALBE
        if "ANGLE GALBE" in desc_upper:
            return "ANGLE_GALBE"
        
        # CUBIQUE seul (seulement si pas précédé de "ANGLE")
        # On a déjà vérifié "ANGLE CUBIQUE" plus haut, donc si on arrive ici,
        # on cherche "CUBIQUE" seul (sans "ANGLE" juste avant)
        if "CUBIQUE" in desc_upper:
            # Vérifier qu'il n'y a pas "ANGLE" dans un contexte proche avant "CUBIQUE"
            cubique_index = desc_upper.find("CUBIQUE")
            # Chercher "ANGLE" dans les 15 caractères avant "CUBIQUE" pour être plus robuste
            before_cubique = desc_upper[max(0, cubique_index - 15):cubique_index]
            # Vérifier aussi qu'on n'a pas déjà détecté "ANGLE CUBIQUE" (sécurité supplémentaire)
            if "ANGLE" not in before_cubique and "ANGLE CUBIQUE" not in desc_upper:
                return "CUBIQUE"
        
        # CYLINDRE
        if "CYLINDRE" in desc_upper:
            return "CYLINDRE"
        
        return None
    
    def extraire_nombre_pieds(self, description: str) -> Optional[int]:
        """
        Extrait le nombre de pieds depuis une description.
        
        Args:
            description: Description de l'article
            
        Returns:
            Nombre de pieds trouvé ou None
        """
        desc_upper = description.upper()
        
        # Chercher "X PIEDS" ou "X PIED" (au singulier ou pluriel)
        match = re.search(r'(\d+)\s*PIEDS?', desc_upper)
        if match:
            return int(match.group(1))
        
        return None
    
    def extraire_texte_apres_type(self, description: str, type_pieds: str) -> str:
        """
        Extrait le texte entre le type de pieds et le premier "+".
        
        Args:
            description: Description complète
            type_pieds: Type de pieds détecté ("ANGLE_CUBIQUE", "CUBIQUE", etc.)
            
        Returns:
            Texte extrait (sans l'espace après le type) ou chaîne vide
        """
        desc_upper = description.upper()
        
        # Convertir le type en texte à chercher
        type_text = type_pieds.replace("_", " ")  # "ANGLE_CUBIQUE" -> "ANGLE CUBIQUE"
        
        # Trouver la position du type dans la description
        type_index = desc_upper.find(type_text)
        if type_index == -1:
            return ""
        
        # Position après le type (en tenant compte de la longueur du type)
        after_type_index = type_index + len(type_text)
        
        # Chercher le premier "+" après le type
        plus_index = desc_upper.find("+", after_type_index)
        
        if plus_index == -1:
            # Pas de "+" trouvé, ne rien retourner
            return ""
        
        # Extraire le texte entre le type et le "+"
        texte_extraite = description[after_type_index:plus_index].strip()
        
        return texte_extraite
    
    def detecter_platine_reunion(self, description: str) -> bool:
        """
        Détecte la présence de "PLATINE DE REUNION" ou "PLATINE DE RÉUNION" dans la description.
        
        Args:
            description: Description de l'article
            
        Returns:
            True si "PLATINE DE REUNION" est détecté, False sinon
        """
        # Normaliser la description (enlever accents, mettre en majuscules)
        desc_normalized = unicodedata.normalize('NFKD', description.upper()).encode('ASCII', 'ignore').decode('ASCII')
        
        # Chercher "PLATINE(S) DE REUNION" (sans accent après normalisation, singulier ou pluriel)
        if "PLATINE DE REUNION" in desc_normalized or "PLATINES DE REUNION" in desc_normalized:
            return True

        # Chercher aussi avec accent dans la description originale (au cas où)
        desc_upper = description.upper()
        if "PLATINE DE RÉUNION" in desc_upper or "PLATINES DE RÉUNION" in desc_upper:
            return True
        
        return False
    
    def detecter_type_patin(self, description: str) -> Optional[str]:
        """
        Détecte le type de patin dans la description.
        
        Args:
            description: Description de l'article
            
        Returns:
            Type de patin détecté: "TEFLON", "CARRELAGE", "FEUTRE", ou None
        """
        desc_upper = description.upper()
        
        # Vérifier d'abord qu'il y a bien "PATIN" ou "PATINS" dans la description
        if "PATIN" not in desc_upper:
            return None
        
        # Ordre de recherche : chercher les types spécifiques avec "PATIN" ou "PATINS"
        # Chercher "PATIN TEFLON" ou "PATINS TEFLON" en priorité
        if "PATIN TEFLON" in desc_upper or "PATINS TEFLON" in desc_upper:
            return "TEFLON"
        
        # Chercher "PATIN CARRELAGE" ou "PATINS CARRELAGE"
        if "PATIN CARRELAGE" in desc_upper or "PATINS CARRELAGE" in desc_upper:
            return "CARRELAGE"
        
        # Chercher "PATIN FEUTRE" ou "PATINS FEUTRE"
        if "PATIN FEUTRE" in desc_upper or "PATINS FEUTRE" in desc_upper:
            return "FEUTRE"
        
        # Fallback : chercher les mots seuls mais proches de "PATIN"
        # Chercher "TEFLON" près de "PATIN" (dans les 20 caractères)
        teflon_index = desc_upper.find("TEFLON")
        if teflon_index != -1:
            patin_index = desc_upper.find("PATIN")
            if patin_index != -1 and abs(teflon_index - patin_index) <= 20:
                return "TEFLON"
        
        carrelage_index = desc_upper.find("CARRELAGE")
        if carrelage_index != -1:
            patin_index = desc_upper.find("PATIN")
            if patin_index != -1 and abs(carrelage_index - patin_index) <= 20:
                return "CARRELAGE"
        
        feutre_index = desc_upper.find("FEUTRE")
        if feutre_index != -1:
            patin_index = desc_upper.find("PATIN")
            if patin_index != -1 and abs(feutre_index - patin_index) <= 20:
                return "FEUTRE"
        
        return None
    
    def detecter_sommier_jumeaux(self, description: str) -> bool:
        """
        Détecte si "JUMEAUX" est présent dans une description de sommier.
        
        Args:
            description: Description de l'article
            
        Returns:
            True si "JUMEAUX" est détecté, False sinon
        """
        desc_upper = description.upper()
        return "JUMEAUX" in desc_upper
    
    def extraire_dimensions_sommier(self, description: str) -> Optional[tuple]:
        """
        Extrait les dimensions au format "159/ 199/ 17" depuis une description.
        
        Args:
            description: Description de l'article
            
        Returns:
            Tuple (valeur1, valeur2, valeur3) ou None si non trouvé
        """
        # Chercher un pattern comme "159/ 199/ 17" ou "159/199/17" ou "159 / 199 / 17"
        # Pattern: nombre(s) / nombre(s) / nombre(s)
        pattern = r'(\d+)\s*/\s*(\d+)\s*/\s*(\d+)'
        match = re.search(pattern, description)
        
        if match:
            try:
                val1 = int(match.group(1))
                val2 = int(match.group(2))
                val3 = int(match.group(3))
                return (val1, val2, val3)
            except ValueError:
                return None
        
        return None
    
    def formater_dimensions_sommier(self, dimensions: tuple) -> str:
        """
        Formate les dimensions en multipliant simplement par 10 (sans arrondi).
        Les dimensions de la structure sont prises telles quelles.
        Ex: (159, 199, 19) → "1590 X 1990 X 190"
        Ex: (179, 199, 19) → "1790 X 1990 X 190"

        Args:
            dimensions: Tuple (valeur1, valeur2, valeur3)

        Returns:
            Chaîne formatée : "1590 X 1990 X 190"
        """
        val1, val2, val3 = dimensions

        # Multiplier par 10 sans arrondi
        val1_mm = int(val1 * 10)
        val2_mm = int(val2 * 10)
        val3_mm = int(val3 * 10)

        # Format : "1590 X 1990 X 190"
        return f"{val1_mm} X {val2_mm} X {val3_mm}"
    
    def extraire_texte_apres_dernier_tiret(self, description: str) -> str:
        """
        Extrait le texte après le dernier "-" dans la description.
        
        Args:
            description: Description complète
            
        Returns:
            Texte après le dernier "-" ou chaîne vide
        """
        # Trouver le dernier "-"
        dernier_tiret_index = description.rfind("-")
        if dernier_tiret_index == -1:
            return ""
        
        # Extraire le texte après le dernier "-"
        texte_apres = description[dernier_tiret_index + 1:].strip()
        return texte_apres
    
    def enlever_dimensions_texte(self, texte: str) -> str:
        """
        Enlève les dimensions du texte (format comme "160/ 84 CM" ou "160/84CM").
        
        Args:
            texte: Texte à nettoyer
            
        Returns:
            Texte sans les dimensions
        """
        # Pattern pour détecter les dimensions : nombre/nombre CM ou nombre/nombreCM
        pattern = r'\d+\s*/\s*\d+\s*CM?'
        texte_sans_dim = re.sub(pattern, '', texte, flags=re.IGNORECASE)
        return texte_sans_dim.strip()
    
    def detecter_finition_structure(self, description: str) -> List[tuple]:
        """
        Détecte toutes les finitions de la structure dans toute la description.
        Le texte écrit dans Excel est extrait après le dernier "-" (sans dimensions).
        
        Args:
            description: Description complète
            
        Returns:
            Liste de tuples (type_finition, texte)
            Types possibles: "PAREMENTEE", "CHENE", "HETRE", "FRENE"
        """
        finitions = []
        
        if not description:
            return finitions
        
        # Chercher dans toute la description (pas seulement après le dernier "-")
        desc_upper = description.upper()
        desc_normalized = unicodedata.normalize('NFKD', desc_upper).encode('ASCII', 'ignore').decode('ASCII')
        
        # Extraire le texte après le dernier "-" pour l'écriture dans Excel
        texte_apres_tiret = self.extraire_texte_apres_dernier_tiret(description)
        texte_clean = self.enlever_dimensions_texte(texte_apres_tiret) if texte_apres_tiret else ""
        
        # Retirer le contexte des lattes pour éviter les faux positifs
        # Ex: "DOUBLES LATTES ( HÊTRE MULTIPLIS )" ne doit pas compter comme finition HÊTRE
        # Gérer les variations d'espaces et d'accents dans les parenthèses
        desc_sans_lattes = re.sub(r'LATTES\s*\(\s*[^)]*\)', 'LATTES', desc_normalized)
        # Gérer aussi le format sans parenthèses: "LATTES HETRE MULTIPLIS" après un tiret
        desc_sans_lattes = re.sub(r'LATTES\s+HETRE\s+MULTIPLIS', 'LATTES', desc_sans_lattes)

        # Détecter PAREMENTÉ (avec ou sans "STRUCTURE" devant)
        # Gérer aussi "PAREMENT EE" ou "PAREMENT E" (espace due à l'extraction PDF)
        if ("PAREMENTEE" in desc_sans_lattes or "PAREMENTE" in desc_sans_lattes
                or "PAREMENT EE" in desc_sans_lattes or "PAREMENT E " in desc_sans_lattes):
            finitions.append(("PAREMENTEE", texte_clean))

        # Détecter STRUCTURE HÊTRE MULTIPLIS → type MULTIPLIS (row 25), pas HÊTRE (row 27)
        if "HETRE MULTIPLIS" in desc_sans_lattes or "STRUCTURE HETRE" in desc_sans_lattes and "MULTIPLIS" in desc_sans_lattes:
            finitions.append(("MULTIPLIS", texte_clean))
        # Détecter HÊTRE seul (ex: "HÊTRE LAQUÉ") → row 27
        elif "HETRE" in desc_sans_lattes:
            finitions.append(("HETRE", texte_clean))

        # Détecter FRÊNE
        if "FRENE" in desc_sans_lattes:
            finitions.append(("FRENE", texte_clean))

        # Détecter CHÊNE
        if "CHENE" in desc_sans_lattes:
            finitions.append(("CHENE", texte_clean))
        
        return finitions
    
    def detecter_telecommande_type(self, description: str) -> Optional[str]:
        """
        Détecte le type de télécommande dans la description.
        
        Args:
            description: Description de l'article
            
        Returns:
            "RADIO" si RADIO/SANS FIL/SS FIL détecté, "STANDARD" sinon, ou None si pas de télécommande
        """
        desc_upper = description.upper()
        
        # Vérifier d'abord qu'il y a "TÉLÉCOMMANDE" ou "TELECOMMANDE"
        if "TÉLÉCOMMANDE" not in desc_upper and "TELECOMMANDE" not in desc_upper:
            return None
        
        # Chercher RADIO, SANS FIL, SS FIL
        if "RADIO" in desc_upper or "SANS FIL" in desc_upper or "SS FIL" in desc_upper:
            return "RADIO"
        
        return "STANDARD"
    
    def ecrire_tete_dosseret_excel(self, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                                  description: str, left_col: str, right_col: str) -> None:
        """
        Détecte et écrit les informations de TETE DROITE AJOUREE, TETE/TÊTE, DOSSERET et CHEVET dans Excel.
        
        Args:
            worksheet: Feuille de calcul
            description: Description de l'article
            left_col: Colonne gauche du bloc (B, D, F, H, J, N, P, R, T, V)
            right_col: Colonne droite du bloc (C, E, G, I, K, O, Q, S, U, W)
        """
        desc_upper = description.upper()
        
        # Détecter TETE DROITE AJOUREE (priorité)
        if desc_upper.startswith("TETE DROITE AJOUREE"):
            # Écrire "X" en B9, D9, F9...
            cell_address_left = f"{left_col}9"
            worksheet[cell_address_left] = "X"
            logger.info(f"Écriture TETE DROITE AJOUREE: {cell_address_left} = X")
            
            # Copier la description complète en C9, E9, G9... (même ligne)
            cell_address_right = f"{right_col}9"
            worksheet[cell_address_right] = description
            logger.info(f"Écriture description TETE: {cell_address_right} = {description[:50]}...")
        
        # Détecter TETE/TÊTE (commence par, sans "DROITE AJOUREE")
        elif desc_upper.startswith("TETE") or desc_upper.startswith("TÊTE"):
            # Écrire "X" en B9, D9, F9...
            cell_address_left = f"{left_col}9"
            worksheet[cell_address_left] = "X"
            logger.info(f"Écriture TETE: {cell_address_left} = X")
            
            # Copier la description complète en C9, E9, G9... (même ligne)
            cell_address_right = f"{right_col}9"
            worksheet[cell_address_right] = description
            logger.info(f"Écriture description TETE: {cell_address_right} = {description[:50]}...")
        
        # Détecter DOSSERET (insensible à la casse) - l'article doit commencer par DOSSERET
        # Exclure les mentions incidentes ("JUMEAUX SUR DOSSERET", "pour dosseret + sommier")
        elif desc_upper.startswith("DOSSERET"):
            # Écrire "X" en B10, D10, F10...
            cell_address_left = f"{left_col}10"
            worksheet[cell_address_left] = "X"
            logger.info(f"Écriture DOSSERET: {cell_address_left} = X")

            # Copier la description nettoyée en C10, E10, G10... (même ligne)
            cell_address_right = f"{right_col}10"
            # Nettoyer: normaliser espaces, supprimer après "BASE SOMMIERS"
            desc_clean = re.sub(r'\s+', ' ', description).strip()
            # Couper après "BASE SOMMIERS" si présent
            idx_base = desc_clean.upper().find("BASE SOMMIERS")
            if idx_base >= 0:
                desc_clean = desc_clean[:idx_base + len("BASE SOMMIERS")]
            # Normaliser les espaces autour des "/" dans les dimensions
            desc_clean = re.sub(r'\s*/\s*', '/', desc_clean)
            # Supprimer les accents pour uniformiser
            desc_clean = unicodedata.normalize('NFKD', desc_clean).encode('ASCII', 'ignore').decode('ASCII')
            worksheet[cell_address_right] = desc_clean
            logger.info(f"Écriture description DOSSERET: {cell_address_right} = {desc_clean[:50]}...")
        
        # Détecter CHEVET, TIROIR ou NICHE
        if "CHEVET" in desc_upper or "TIROIR" in desc_upper or "NICHE" in desc_upper:
            # Écrire "X" en B13, D13, F13...
            cell_address_left = f"{left_col}13"
            worksheet[cell_address_left] = "X"
            logger.info(f"Écriture CHEVET/TIROIR: {cell_address_left} = X")

            # Copier la description complète en C13, E13, G13... (même ligne)
            cell_address_right = f"{right_col}13"
            worksheet[cell_address_right] = description
            logger.info(f"Écriture description CHEVET/TIROIR: {cell_address_right} = {description[:50]}...")
    
    def detecter_type_sommier_fixe_manuel_motorise(self, description: str) -> Optional[str]:
        """
        Détecte le type de sommier détaillé depuis une description.

        Returns:
            "FIXE", "TPR", "TT_TENON", "TT_TPR", ou None
        """
        if not description:
            return None

        desc_upper = description.upper()
        desc_norm = unicodedata.normalize('NFKD', desc_upper).encode('ASCII', 'ignore').decode('ASCII')
        # Normaliser les espaces multiples (PDF extraction peut créer "MOTORIS EE" ou "TELESCOP IQUE")
        desc_norm = re.sub(r'\s+', ' ', desc_norm)

        # 1. FIXE (pas de relaxation) - normaliser les tirets pour gérer "JUMEAUX - FIXE"
        desc_fixe = desc_norm.replace(" - ", " ").replace("- ", " ").replace(" -", " ")
        if ("SOMMIER FIXE" in desc_fixe or "SOMMIERS JUMEAUX FIXE" in desc_fixe
                or "SOMMIERS FIXE" in desc_fixe or "JUMEAUX FIXE" in desc_fixe):
            return "FIXE"

        # 2. Relaxation motorisée télescopique → TT embout sur TENON
        has_telescopique = "TELESCOPIQUE" in desc_norm or "TELESCOP" in desc_norm
        has_motorise = "MOTORISEE" in desc_norm or "MOTORISE" in desc_norm or "MOTORIS" in desc_norm
        if has_telescopique and has_motorise:
            return "TT_TENON"

        # 3. Relaxation manuelle → TPR
        if "RELAXATION MANUELLE" in desc_norm or "MANUEL" in desc_norm:
            return "TPR"

        # 4. Relaxation motorisée (non télescopique) → TT embout TPR
        if has_motorise or "RELAXATION" in desc_norm:
            return "TT_TPR"

        return None
    
    def ecrire_type_sommier_fixe_manuel_motorise_excel(self, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                                                       description: str, left_col: str,
                                                       right_col: str = None) -> None:
        """
        Écrit un "X" dans la cellule Excel appropriée selon le type de sommier détecté.
        Mapping template :
          row 18 = FIXE, row 19 = TPR, row 20 = TT embout TPR, row 21 = TT embout sur TENON

        Pour les FIXE, écrit aussi "LATTES DESSUS" ou "LATTES DESSOUS" dans right_col row 18.
        """
        type_sommier = self.detecter_type_sommier_fixe_manuel_motorise(description)

        if not type_sommier:
            return

        # Mapping des types vers les lignes Excel du template
        ligne_mapping = {
            "FIXE": 18,       # row 18 = FIXE
            "TPR": 19,        # row 19 = TPR
            "TT_TPR": 20,     # row 20 = TT embout TPR
            "TT_TENON": 21,   # row 21 = TT embout sur TENON
        }

        ligne = ligne_mapping.get(type_sommier)
        if ligne:
            cell_address = f"{left_col}{ligne}"
            worksheet[cell_address] = "X"
            logger.info(f"Écriture type sommier: {cell_address} = X ({type_sommier})")

            # Pour FIXE : ajouter la précision LATTES DESSUS / LATTES DESSOUS
            if type_sommier == "FIXE" and right_col:
                desc_upper = description.upper()
                if "LATTES DESSUS" in desc_upper:
                    worksheet[f"{right_col}{ligne}"] = "LATTES DESSUS"
                    logger.info(f"Écriture FIXE détail: {right_col}{ligne} = LATTES DESSUS")
                elif "LATTES DESSOUS" in desc_upper:
                    worksheet[f"{right_col}{ligne}"] = "LATTES DESSOUS"
                    logger.info(f"Écriture FIXE détail: {right_col}{ligne} = LATTES DESSOUS")

        # Pour les types TT (motorisés), marquer aussi row 23 (RF Noire + Torche)
        # si la télécommande RF est déjà gérée par ecrire_telecommande_excel
        # Pour TPR, marquer aussi TT_TENON si c'est un 3 plis (relaxation manuelle = TPR + TT)
        if type_sommier == "TPR":
            # Les sommiers relaxation manuelle ont aussi l'embout sur TENON
            cell_tt = f"{left_col}21"
            worksheet[cell_tt] = "X"
            logger.info(f"Écriture TPR + TT TENON: {cell_tt} = X")
    
    def ecrire_sommier_excel(self, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                            description: str, left_col: str, right_col: str) -> None:
        """
        Détecte et écrit les informations de sommier dans Excel.
        - Détecte JUMEAUX : écrit "X" en B11 ou B12
        - Extrait et formate les dimensions : écrit la chaîne formatée
        
        Args:
            worksheet: Feuille de calcul
            description: Description de l'article commençant par "SOMMIER"
            left_col: Colonne gauche du bloc (B, D, F, H, J, N, P, R, T, V)
            right_col: Colonne droite du bloc (C, E, G, I, K, O, Q, S, U, W)
        """
        desc_upper = description.upper()
        
        # Vérifier que la description commence par "SOMMIER" ou "SOMMIERS"
        if not (desc_upper.startswith("SOMMIER") or desc_upper.startswith("SOMMIERS")):
            return
        
        # Détecter JUMEAUX
        is_jumeaux = self.detecter_sommier_jumeaux(description)
        
        if is_jumeaux:
            # Écrire "X" en B11, D11, F11...
            cell_address = f"{left_col}11"
            worksheet[cell_address] = "X"
            logger.info(f"Écriture sommier jumeaux: {cell_address} = X")
        else:
            # Écrire "X" en B12, D12, F12...
            cell_address = f"{left_col}12"
            worksheet[cell_address] = "X"
            logger.info(f"Écriture sommier standard: {cell_address} = X")
        
        # Note: Les dimensions sont écrites dans write_config_to_block() selon la présence de pieds
        # (C7 si pieds détectés, C6 sinon). C13 est réservé pour CHEVET.
    
    def ecrire_finition_structure_excel(self, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                                       description: str, left_col: str, right_col: str) -> None:
        """
        Détecte et écrit toutes les finitions de la structure dans Excel.
        
        Args:
            worksheet: Feuille de calcul
            description: Description de l'article
            left_col: Colonne gauche du bloc (B, D, F, H, J, N, P, R, T, V)
            right_col: Colonne droite du bloc (C, E, G, I, K, O, Q, S, U, W)
        """
        finitions = self.detecter_finition_structure(description)
        
        if not finitions:
            return
        
        # Mapping des types vers les lignes Excel
        ligne_mapping = {
            "MULTIPLIS": 25,   # B25, D25, F25...
            "PAREMENTEE": 26,  # B26, D26, F26...
            "HETRE": 27,       # B27, D27, F27...
            "FRENE": 28,       # B28, D28, F28...
            "CHENE": 29,       # B29, D29, F29...
        }
        
        # Écrire toutes les finitions détectées
        for type_finition, texte in finitions:
            ligne = ligne_mapping.get(type_finition)
            if ligne:
                # Écrire "X" dans la colonne gauche
                cell_address_left = f"{left_col}{ligne}"
                worksheet[cell_address_left] = "X"
                logger.info(f"Écriture finition structure: {cell_address_left} = X ({type_finition})")
                
                # Écrire le texte (sans dimensions) dans la colonne droite
                if texte:
                    cell_address_right = f"{right_col}{ligne}"
                    worksheet[cell_address_right] = texte
                    logger.info(f"Écriture texte finition: {cell_address_right} = {texte[:50]}...")
    
    def ecrire_telecommande_excel(self, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                                  description: str, left_col: str) -> None:
        """
        Détecte et écrit le type de télécommande dans Excel.
        
        Args:
            worksheet: Feuille de calcul
            description: Description de l'article
            left_col: Colonne gauche du bloc (B, D, F, H, J, N, P, R, T, V)
        """
        type_telecommande = self.detecter_telecommande_type(description)
        
        if not type_telecommande:
            return
        
        # Mapping des types vers les lignes Excel
        if type_telecommande == "RADIO":
            # RADIO/SANS FIL/SS FIL → B23, D23, F23...
            cell_address = f"{left_col}23"
            worksheet[cell_address] = "X"
            logger.info(f"Écriture télécommande radio: {cell_address} = X")
        else:
            # STANDARD → B22, D22, F22...
            cell_address = f"{left_col}22"
            worksheet[cell_address] = "X"
            logger.info(f"Écriture télécommande standard: {cell_address} = X")
    
    def detecter_lot_rampes(self, description: str) -> bool:
        """
        Détecte "LOT DE 2 RAMPES" dans la description.
        
        Args:
            description: Description de l'article
            
        Returns:
            True si "LOT DE 2 RAMPES" est détecté, False sinon
        """
        if not description:
            return False
        
        desc_upper = description.upper()
        return "LOT DE 2 RAMPES" in desc_upper
    
    def detecter_lattes_dessus(self, description: str) -> bool:
        """
        Détecte "LATTES DESSUS" dans la description.
        
        Args:
            description: Description de l'article
            
        Returns:
            True si "LATTES DESSUS" est détecté, False sinon
        """
        if not description:
            return False
        
        desc_upper = description.upper()
        return "LATTES DESSUS" in desc_upper
    
    def calculer_dimensions_modifiees(self, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                                     description: str, left_col: str, right_col: str,
                                     dimensions_originales: Optional[tuple] = None) -> Optional[str]:
        """
        Calcule les dimensions jumeaux selon les règles basées sur le type de sommier et la finition.

        Règles observées depuis le fichier corrigé :
        - Motorisé/Manuel (TT/TPR) :
          * PAREMENTÉE/CHÊNE : largeur -10, hauteur fixe 140
          * HÊTRE/FRÊNE/MULTIPLIS : largeur -6, hauteur fixe 140
        - FIXE :
          * Pas de déduction, dimensions conservées telles quelles
          * Hauteur décimale formatée avec virgule (ex: 11,5)

        La longueur est toujours la longueur originale * 10.
        """
        if not dimensions_originales:
            return None

        largeur_cm, longueur_cm, hauteur_cm = dimensions_originales

        # Dimensions de base en mm (sans arrondi)
        largeur_mm = int(largeur_cm * 10)
        longueur_mm = int(longueur_cm * 10)
        hauteur_mm = int(hauteur_cm * 10)

        # Détecter le type de sommier
        type_sommier = self.detecter_type_sommier_fixe_manuel_motorise(description)

        # Détecter la finition
        finitions = self.detecter_finition_structure(description)
        finition_types = [f[0] for f in finitions]

        # Détecter LATTES DESSUS
        has_lattes_dessus = self.detecter_lattes_dessus(description)

        # Appliquer les règles de calcul des jumeaux
        nouvelle_largeur = largeur_mm
        nouvelle_longueur = longueur_mm
        nouvelle_hauteur = hauteur_mm

        if type_sommier in ["TPR", "TT_TENON", "TT_TPR"]:
            # Motorisé ou Relaxation manuelle : hauteur fixe 140mm
            nouvelle_hauteur = 140
            if "PAREMENTEE" in finition_types or "CHENE" in finition_types:
                nouvelle_largeur = largeur_mm - 10
            elif "HETRE" in finition_types or "FRENE" in finition_types or "MULTIPLIS" in finition_types:
                nouvelle_largeur = largeur_mm - 6

        elif type_sommier == "FIXE":
            # FIXE : pas de déduction sur la largeur, dimensions conservées telles quelles
            pass

        # Formater la chaîne de résultat
        # Utiliser virgule pour hauteur si c'est un nombre décimal (ex: 11,5)
        if hauteur_cm != int(hauteur_cm) and type_sommier == "FIXE":
            hauteur_str = str(hauteur_cm).replace('.', ',')
            resultat = f"{nouvelle_largeur} X {nouvelle_longueur} X {hauteur_str}"
        else:
            resultat = f"{nouvelle_largeur} X {nouvelle_longueur} X {nouvelle_hauteur}"

        logger.info(f"Dimensions jumeaux calculées: {resultat} (type: {type_sommier}, finitions: {finition_types}, lattes_dessus: {has_lattes_dessus})")

        return resultat
    
    def ecrire_lot_rampes_excel(self, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                               description: str, left_col: str) -> None:
        """
        Détecte "LOT DE 2 RAMPES" et écrit "X" en B54, D54, F54...
        
        Args:
            worksheet: Feuille de calcul
            description: Description de l'article
            left_col: Colonne gauche du bloc (B, D, F, H, J, N, P, R, T, V)
        """
        if self.detecter_lot_rampes(description):
            cell_address = f"{left_col}54"
            worksheet[cell_address] = "X"
            logger.info(f"Écriture lot rampes: {cell_address} = X")
    
    def ecrire_type_pieds_excel(self, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                               description: str, left_col: str, right_col: str) -> None:
        """
        Écrit un "X" dans la cellule Excel appropriée selon le type de pieds détecté,
        et écrit le nombre de pieds + texte dans la colonne droite à la ligne + 1.
        Détecte aussi PLATINE DE REUNION et les types de patins.
        
        Args:
            worksheet: Feuille de calcul
            description: Description de l'article contenant les pieds
            left_col: Colonne gauche du bloc (B, D, F, H, J, N, P, R, T, V)
            right_col: Colonne droite du bloc (C, E, G, I, K, O, Q, S, U, W)
        """
        # Détecter le type de pieds
        type_pieds = self.detecter_type_pieds(description)
        
        if type_pieds:
            # Mapping des types vers les lignes Excel
            ligne_mapping = {
                "ANGLE_CUBIQUE": 33,  # B33, D33, F33...
                "ANGLE_DROIT": 35,    # B35, D35, F35...
                "ANGLE_GALBE": 34,    # B34, D34, F34...
                "CUBIQUE": 36,        # B36, D36, F36...
                "CYLINDRE": 37,       # B37, D37, F37...
            }
            
            ligne = ligne_mapping.get(type_pieds)
            if ligne:
                # Écrire "X" dans la colonne gauche
                cell_address_left = f"{left_col}{ligne}"
                worksheet[cell_address_left] = "X"
                logger.info(f"Écriture type pieds: {cell_address_left} = X ({type_pieds})")
                
                # Extraire le nombre de pieds
                nb_pieds = self.extraire_nombre_pieds(description)
                
                # Extraire le texte entre le type et le premier "+"
                texte_apres_type = self.extraire_texte_apres_type(description, type_pieds)
                
                # Écrire dans la colonne droite sur la même ligne que le "X" seulement si on a un nombre ET un texte
                if nb_pieds is not None and texte_apres_type:
                    cell_address_right = f"{right_col}{ligne}"
                    valeur = f"{nb_pieds}x {texte_apres_type}"
                    worksheet[cell_address_right] = valeur
                    logger.info(f"Écriture type pieds: {cell_address_right} = {valeur}")
        
        # Détecter et écrire PLATINE DE REUNION (indépendamment du type de pieds)
        if self.detecter_platine_reunion(description):
            cell_address_platine = f"{left_col}39"
            worksheet[cell_address_platine] = "X"
            logger.info(f"Écriture platine réunion: {cell_address_platine} = X")

        # Détecter et écrire PIEDS CENTRAUX / PIED CENTRAL
        desc_pieds_upper = description.upper()
        if "PIED CENTRAL" in desc_pieds_upper or "PIEDS CENTRAUX" in desc_pieds_upper:
            cell_address_centraux = f"{left_col}40"
            worksheet[cell_address_centraux] = "X"
            logger.info(f"Écriture pieds centraux: {cell_address_centraux} = X")
            # Extraire le nombre de pieds centraux si mentionné
            match_nb = re.search(r'(\d+)\s*PIED(?:S)?\s*CENTR', desc_pieds_upper)
            if match_nb:
                nb_centraux = int(match_nb.group(1))
                worksheet[f"{right_col}40"] = nb_centraux
                logger.info(f"Nombre pieds centraux: {right_col}40 = {nb_centraux}")

        # Détecter et écrire le type de patin (indépendamment du type de pieds)
        type_patin = self.detecter_type_patin(description)
        if type_patin:
            ligne_patin_mapping = {
                "FEUTRE": 41,      # B41, D41, F41...
                "CARRELAGE": 42,   # B42, D42, F42...
                "TEFLON": 43,      # B43, D43, F43...
            }
            ligne_patin = ligne_patin_mapping.get(type_patin)
            if ligne_patin:
                cell_address_patin = f"{left_col}{ligne_patin}"
                worksheet[cell_address_patin] = "X"
                logger.info(f"Écriture type patin: {cell_address_patin} = X ({type_patin})")
    
    def write_config_to_block(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                             config_json: Dict, left_col: str, right_col: str) -> None:
        """
        Écrit une configuration sommier dans un bloc spécifique
        
        Args:
            worksheet: Feuille de calcul
            config_json: Configuration JSON du sommier
            left_col: Colonne de gauche
            right_col: Colonne de droite
        """
        self._clear_block(worksheet, left_col, right_col)
        # Vérifier si PAREMENTEE est détecté pour forcer MULTIPLIS à vide dans Excel
        description_sommier = config_json.get("description", "") or config_json.get("article_description", "")
        has_paremente = False
        if description_sommier:
            desc_upper = description_sommier.upper()
            desc_normalized = description_sommier.upper().replace('É', 'E').replace('Ê', 'E')
            if ("STRUCTURE PAREMENTEE" in desc_normalized or 
                "STRUCTURE PAREMENTÉE" in desc_upper or 
                "PAREMENTE" in desc_normalized or
                "PAREMENTÉE" in desc_upper):
                has_paremente = True
                logger.info(f"PAREMENTEE détecté dans write_config_to_block: {description_sommier[:100]}")
        has_sommier_article = False
        
        # Mapper les données vers les cellules
        cell_mappings = self.map_json_to_cells(config_json, left_col, right_col)
        
        # Si PAREMENTEE est détecté, forcer les cellules MULTIPLIS à vide dans Excel
        if has_paremente:
            # Les champs MULTIPLIS sont mappés vers right_col}76, right_col}79, right_col}80
            # Forcer ces cellules à vide même si elles sont dans cell_mappings
            multiplis_cells = [
                f"{right_col}76",  # Finition_Multiplis_D76
                f"{right_col}79",  # Finition_Multiplis_TV_D79
                f"{right_col}80",  # Finition_Multiplis_L_D80
            ]
            for cell_addr in multiplis_cells:
                if cell_addr in cell_mappings:
                    cell_mappings[cell_addr] = ''
                    logger.info(f"MULTIPLIS forcé à vide dans Excel (write_config_to_block): {cell_addr}")
                # Aussi forcer directement dans Excel pour être sûr
                worksheet[cell_addr] = ''
                logger.info(f"MULTIPLIS forcé à vide directement dans Excel: {cell_addr}")
        
        # Écrire les données
        for cell_address, value in cell_mappings.items():
            # Ne pas écrire les valeurs vides (sauf si explicitement forcé à vide)
            if value:
                # Simplifier le nom client : extraire uniquement le nom de famille
                if cell_address.endswith("1") and cell_address[:-1] in [right_col]:
                    value = self.extraire_nom_famille(value)

                # Convertir les valeurs numériques (numéro de commande)
                if isinstance(value, str) and value.isdigit():
                    value = int(value)

                worksheet[cell_address] = value
                logger.info(f"Écriture: {cell_address} = {value}")

                # Appliquer l'alignement
                json_key = next((k for k, v in cell_mappings.items() if v == value), "")
                self.apply_cell_alignment(worksheet, cell_address, json_key)
        
        # Écrire les champs d'opération dans la colonne gauche du bloc (left_col)
        # B56, D56, F56... pour emporte_client_C57
        if config_json.get("emporte_client_C57") == "X":
            cell_address = f"{left_col}56"
            worksheet[cell_address] = "X"
            logger.info(f"Écriture opération: {cell_address} = X (emporte_client_C57)")

        # B57, D57, F57... pour fourgon_C58
        if config_json.get("fourgon_C58") == "X":
            cell_address = f"{left_col}57"
            worksheet[cell_address] = "X"
            logger.info(f"Écriture opération: {cell_address} = X (fourgon_C58)")

        # B58, D58, F58... pour transporteur_C59
        if config_json.get("transporteur_C59") == "X":
            cell_address = f"{left_col}58"
            worksheet[cell_address] = "X"
            logger.info(f"Écriture opération: {cell_address} = X (transporteur_C59)")
        
        # Écrire les champs de pieds dans la colonne gauche du bloc (left_col)
        # B40, D40, F40... pour Pieds_Centraux_D72
        if config_json.get("Pieds_Centraux_D72") == "X":
            cell_address = f"{left_col}40"
            worksheet[cell_address] = "X"
            logger.info(f"Écriture pieds: {cell_address} = X (Pieds_Centraux_D72)")
        
        # B46, D46, F46... pour Butees_Pieds_D61
        if config_json.get("Butees_Pieds_D61") == "X":
            cell_address = f"{left_col}46"
            worksheet[cell_address] = "X"
            logger.info(f"Écriture pieds: {cell_address} = X (Butees_Pieds_D61)")
        
        # NOTE: Sommier_Pieds_D50 n'est plus écrit nulle part
        
        # Détecter et écrire les types de pieds depuis les articles
        # Chercher dans plusieurs emplacements possibles pour les descriptions de pieds
        
        # Détecter si des pieds sont présents dans la commande
        has_pieds = False
        dimensions_sommier = None
        
        # 1. Chercher dans les articles directement dans config_json
        # Utiliser un set pour tracker ce qui a déjà été écrit dans ce bloc pour éviter les doublons
        articles_traites = set()
        articles = config_json.get("articles", [])
        logger.info(f"write_config_to_block: {len(articles) if isinstance(articles, list) else 0} articles, left={left_col}, right={right_col}")
        if isinstance(articles, list):
            for article in articles:
                if isinstance(article, dict):
                    description = article.get("description", "")
                    if not description:
                        continue
                    # Nettoyer le boilerplate PDF de page 2
                    description = _strip_pdf_boilerplate(description)
                    
                    # Créer un identifiant unique pour cet article (pour éviter les doublons)
                    article_id = f"{article.get('type', '')}_{description[:50]}"
                    if article_id in articles_traites:
                        continue  # Déjà traité
                    articles_traites.add(article_id)
                    
                    # Détecter si c'est un article de pieds
                    if self.detecter_type_pieds(description) or "PIEDS" in description.upper():
                        has_pieds = True
                    
                    # Détecter et écrire TETE DROITE AJOUREE, TETE/TÊTE, DOSSERET et CHEVET
                    # (seulement si ce n'est pas déjà écrit dans ce bloc)
                    self.ecrire_tete_dosseret_excel(worksheet, description, left_col, right_col)
                    
                    # Détecter et écrire la finition de la structure
                    # (seulement si ce n'est pas déjà écrit dans ce bloc)
                    self.ecrire_finition_structure_excel(worksheet, description, left_col, right_col)
                    
                    # Détecter et écrire le type de télécommande
                    # (seulement si ce n'est pas déjà écrit dans ce bloc)
                    self.ecrire_telecommande_excel(worksheet, description, left_col)
                    
                    # Détecter et écrire les informations de sommier (JUMEAUX, dimensions)
                    # (seulement pour les articles de sommier, pas pour tous les articles)
                    # Gérer aussi "SOMMIER S" (espace due à l'extraction PDF) et le type LLM "sommier"
                    desc_up = description.upper().lstrip()
                    is_sommier_article = (
                        desc_up.startswith("SOMMIER") or
                        desc_up.startswith("SOMMIERS") or
                        article.get("type", "").lower() == "sommier"
                    )
                    if is_sommier_article:
                        has_sommier_article = True
                        self.ecrire_sommier_excel(worksheet, description, left_col, right_col)

                        # Détecter et écrire le type de sommier (FIXE, MANUEL, MOTORISE)
                        self.ecrire_type_sommier_fixe_manuel_motorise_excel(worksheet, description, left_col, right_col)

                        # Extraire les dimensions du sommier si présent
                        dims = self.extraire_dimensions_sommier(description)
                        if dims:
                            dimensions_sommier = self.formater_dimensions_sommier(dims)
                        logger.info(f"Article sommier détecté: desc={description[:80]}..., dims={dims}")
                    
                    # Détecter et écrire le type de pieds
                    self.ecrire_type_pieds_excel(worksheet, description, left_col, right_col)

                    # Détecter et écrire LOT DE 2 RAMPES
                    self.ecrire_lot_rampes_excel(worksheet, description, left_col)

                    # Détecter MÉTRAGE PVC ou TISSU → écrire en row 26 (PAREMENTÉ) si PAREMENTÉ déjà détecté
                    desc_metrage = description.upper()
                    if "METRAGE" in desc_metrage or "MÉTRAGE" in desc_metrage:
                        # Extraire le type PVC/TISSU et la référence
                        match_pvc = re.search(r'(?:METRAGE|MÉTRAGE)\s+(PVC|TISSU)\s+(.+?)(?:\s*\d+\s*/\s*\d+|\s*$)', description, re.IGNORECASE)
                        if match_pvc:
                            # Vérifier si PAREMENTÉE est déjà cochée en row 26
                            cell_pare = f"{left_col}26"
                            if worksheet[cell_pare].value == "X":
                                cell_info = f"{right_col}26"
                                ref_text = match_pvc.group(2).strip()
                                # Nettoyer: supprimer tirets, parenthèses de fin
                                ref_text = ref_text.replace('-', '').strip()
                                ref_text = re.sub(r'\s*\(.*?\)\s*$', '', ref_text).strip()
                                ref_text = re.sub(r'\s+', ' ', ref_text)  # normaliser espaces
                                info_text = f"{match_pvc.group(1).upper()} {ref_text}"
                                worksheet[cell_info] = info_text
                                logger.info(f"Écriture métrage sur PAREMENTÉ: {cell_info} = {info_text}")
        
        # 2. Chercher dans la description du sommier lui-même (peut contenir des infos sur les pieds)
        # Skiper si des articles individuels ont déjà été traités (évite les doublons/écrasements)
        articles_processed = isinstance(articles, list) and len(articles) > 1
        description_sommier = config_json.get("description", "") or config_json.get("article_description", "")
        if description_sommier and not articles_processed:
            description_sommier = _strip_pdf_boilerplate(description_sommier)
            if True:
                # Détecter si c'est un article de pieds
                if self.detecter_type_pieds(description_sommier) or "PIEDS" in description_sommier.upper():
                    has_pieds = True
                
                # Détecter et écrire TETE DROITE AJOUREE, TETE/TÊTE, DOSSERET et CHEVET
                self.ecrire_tete_dosseret_excel(worksheet, description_sommier, left_col, right_col)
                
                # Détecter et écrire la finition de la structure
                self.ecrire_finition_structure_excel(worksheet, description_sommier, left_col, right_col)
                
                # Détecter et écrire le type de télécommande
                self.ecrire_telecommande_excel(worksheet, description_sommier, left_col)
                
                # Détecter et écrire les informations de sommier (JUMEAUX, dimensions)
                # (toujours traiter la description du sommier même si déjà dans les articles)
                self.ecrire_sommier_excel(worksheet, description_sommier, left_col, right_col)
                if description_sommier.upper().startswith("SOMMIER") or description_sommier.upper().startswith("SOMMIERS"):
                    has_sommier_article = True
                
                # Détecter et écrire le type de sommier (FIXE, MANUEL, MOTORISE)
                self.ecrire_type_sommier_fixe_manuel_motorise_excel(worksheet, description_sommier, left_col, right_col)
                
                # Extraire les dimensions du sommier si présent
                if description_sommier.upper().startswith("SOMMIER") or description_sommier.upper().startswith("SOMMIERS"):
                    dims = self.extraire_dimensions_sommier(description_sommier)
                    if dims:
                        dimensions_sommier = self.formater_dimensions_sommier(dims)
                
                self.ecrire_type_pieds_excel(worksheet, description_sommier, left_col, right_col)
                
                # Détecter et écrire LOT DE 2 RAMPES
                self.ecrire_lot_rampes_excel(worksheet, description_sommier, left_col)
            else:
                # La description a déjà été traitée, mais on doit quand même extraire les dimensions du sommier
                if description_sommier.upper().startswith("SOMMIER") or description_sommier.upper().startswith("SOMMIERS"):
                    dims = self.extraire_dimensions_sommier(description_sommier)
                    if dims:
                        dimensions_sommier = self.formater_dimensions_sommier(dims)
        
        # 3. Chercher dans les articles de pieds segmentés (si disponibles)
        pieds_segmentes = config_json.get("pieds_segmentes", {})
        if isinstance(pieds_segmentes, dict) and pieds_segmentes:
            has_pieds = True
            type_pied = pieds_segmentes.get("type_pied", "")
            if type_pied:
                # Reconstruire une description pour la détection
                desc_pieds = f"{type_pied}"
                self.ecrire_type_pieds_excel(worksheet, desc_pieds, left_col, right_col)
        
        # Chercher les dimensions LITERIE dans les articles (format "LITERIE ... 160/200/64 CM")
        # Pour row 7, on utilise les dimensions LITERIE (taille du lit), pas les dimensions structure
        dimensions_literie = None
        if isinstance(articles, list):
            for article in articles:
                if isinstance(article, dict):
                    desc = article.get("description", "")
                    if desc and "LITERIE" in desc.upper():
                        dims_lit = self.extraire_dimensions_sommier(desc)
                        if dims_lit:
                            # Prendre largeur/longueur LITERIE mais hauteur depuis SOMMIER structure si disponible
                            hauteur_sommier = None
                            # Chercher d'abord dans les articles de type sommier (plus fiable)
                            if isinstance(articles, list):
                                for art in articles:
                                    if isinstance(art, dict):
                                        d = art.get("description", "")
                                        d_up = d.upper().lstrip() if d else ""
                                        if d_up.startswith("SOMMIER") or art.get("type", "").lower() == "sommier":
                                            d = _strip_pdf_boilerplate(d)
                                            ds = self.extraire_dimensions_sommier(d)
                                            if ds:
                                                hauteur_sommier = ds[2]
                                                break
                            dimensions_literie = self.formater_dimensions_sommier((dims_lit[0], dims_lit[1], hauteur_sommier if hauteur_sommier else dims_lit[2]))
                            logger.info(f"Dimensions LITERIE trouvées: {dims_lit} → {dimensions_literie}")
                            break
        # Si pas de dimensions LITERIE, aussi chercher dans description config_json
        if not dimensions_literie:
            desc_literie = config_json.get("dimension_literie", "") or config_json.get("description_literie", "")
            if desc_literie:
                dims_lit = self.extraire_dimensions_sommier(desc_literie)
                if dims_lit:
                    dimensions_literie = self.formater_dimensions_sommier(dims_lit)

        # Utiliser dimensions LITERIE pour row 6/7 si disponibles, sinon fallback sur dimensions SOMMIER
        dimensions_pour_row = dimensions_literie or dimensions_sommier

        # Écrire les dimensions du sommier selon la présence de pieds
        dimensions_originales_tuple = None
        if dimensions_pour_row:
            if has_pieds:
                # Si des pieds sont trouvés : écrire en C7 et "X" en B7
                cell_address_left = f"{left_col}7"
                cell_address_right = f"{right_col}7"
                worksheet[cell_address_left] = "X"
                worksheet[cell_address_right] = dimensions_pour_row
                logger.info(f"Écriture dimensions avec pieds: {cell_address_left} = X, {cell_address_right} = {dimensions_pour_row}")
            else:
                # Sinon : écrire en C6 et "X" en B6
                cell_address_left = f"{left_col}6"
                cell_address_right = f"{right_col}6"
                worksheet[cell_address_left] = "X"
                worksheet[cell_address_right] = dimensions_pour_row
                logger.info(f"Écriture dimensions sans pieds: {cell_address_left} = X, {cell_address_right} = {dimensions_pour_row}")
            
            # Extraire les dimensions originales (STRUCTURE) pour le calcul des dimensions modifiées
            # Chercher d'abord dans les articles SOMMIER spécifiques (éviter les dims LITERIE)
            if isinstance(articles, list):
                for article in articles:
                    if isinstance(article, dict):
                        desc = article.get("description", "")
                        desc_up = desc.upper().lstrip() if desc else ""
                        if desc_up.startswith("SOMMIER") or article.get("type", "").lower() == "sommier":
                            desc = _strip_pdf_boilerplate(desc)
                            dims = self.extraire_dimensions_sommier(desc)
                            if dims:
                                dimensions_originales_tuple = dims
                                break
            # Fallback: description combinée
            if not dimensions_originales_tuple:
                desc_combined = config_json.get("description", "") or config_json.get("article_description", "")
                if desc_combined:
                    desc_combined = _strip_pdf_boilerplate(desc_combined)
                    dims = self.extraire_dimensions_sommier(desc_combined)
                    if dims:
                        dimensions_originales_tuple = dims

        # Calculer et écrire les dimensions modifiées si applicable
        # Chercher d'abord dans les articles de type sommier (éviter la description combinée)
        description_sommier = None
        if isinstance(articles, list):
            for article in articles:
                if isinstance(article, dict):
                    desc = article.get("description", "")
                    desc_up = desc.upper().lstrip() if desc else ""
                    if desc_up.startswith("SOMMIER") or article.get("type", "").lower() == "sommier":
                        description_sommier = _strip_pdf_boilerplate(desc)
                        break
        # Fallback: description combinée
        if not description_sommier:
            description_sommier = config_json.get("description", "") or config_json.get("article_description", "")
            if description_sommier:
                description_sommier = _strip_pdf_boilerplate(description_sommier)
        if description_sommier and dimensions_originales_tuple:
            dimensions_modifiees = self.calculer_dimensions_modifiees(
                worksheet, description_sommier, left_col, right_col, dimensions_originales_tuple
            )
            if dimensions_modifiees:
                # Écrire les dimensions modifiées en C11 ou C12 selon la coche en B11 ou B12
                cell_b11 = f"{left_col}11"
                cell_b12 = f"{left_col}12"
                
                if worksheet[cell_b11].value == "X":
                    # JUMEAUX détecté -> écrire en C11
                    cell_address_dim_mod = f"{right_col}11"
                    worksheet[cell_address_dim_mod] = dimensions_modifiees
                    logger.info(f"Écriture dimensions modifiées (JUMEAUX): {cell_address_dim_mod} = {dimensions_modifiees}")
                elif worksheet[cell_b12].value == "X":
                    # Pas de JUMEAUX -> écrire en C12
                    cell_address_dim_mod = f"{right_col}12"
                    worksheet[cell_address_dim_mod] = dimensions_modifiees
                    logger.info(f"Écriture dimensions modifiées (standard): {cell_address_dim_mod} = {dimensions_modifiees}")
                else:
                    # Par défaut, essayer de détecter depuis la description
                    is_jumeaux = self.detecter_sommier_jumeaux(description_sommier)
                    if is_jumeaux:
                        cell_address_dim_mod = f"{right_col}11"
                    else:
                        cell_address_dim_mod = f"{right_col}12"
                    worksheet[cell_address_dim_mod] = dimensions_modifiees
                    logger.info(f"Écriture dimensions modifiées (détection): {cell_address_dim_mod} = {dimensions_modifiees}")
        
        
        cell_address_31 = f"{left_col}31"
        if has_sommier_article or config_json.get("type_article", "sommier") == "sommier":
            worksheet[cell_address_31] = "X"
            logger.info(f"Écriture sommier détecté: {cell_address_31} = X (ligne 31)")
        else:
            worksheet[cell_address_31] = None
            logger.info(f"Bloc accessoires uniquement: {cell_address_31} laissé vide")

        # Centrer toutes les cellules du bloc
        self.center_block_cells(worksheet, left_col, right_col)
        self.total_written_cases += 1

    def create_new_file(self, semaine: str, id_fichier: str) -> openpyxl.Workbook:
        """
        Crée un nouveau fichier Excel sommier à partir du template
        
        Args:
            semaine: Code semaine
            id_fichier: ID du fichier
            
        Returns:
            Nouveau workbook
        """
        self.current_file_index += 1
        self.current_case_count = 0
        
        logger.info(f"Création nouveau fichier sommier: index {self.current_file_index}")
        print(f"DEBUG SOMMIER: create_new_file appelée avec semaine={semaine}")
        
        # Charge le template pour le nouveau fichier
        new_workbook = self.load_template()
        new_worksheet = new_workbook.active
        
        # Met à jour les numéros de cas pour ce nouveau fichier
        self.update_case_numbers(new_worksheet, self.current_file_index)
        
        # Ajouter le numéro de semaine en A4
        try:
            # Extraire le numéro de semaine (enlever le 'S' du début)
            numero_semaine = semaine.replace('S', '') if semaine.startswith('S') else semaine
            semaine_text = f"sem. {numero_semaine}"
            
            # Forcer l'écriture en A4, même si cellules fusionnées
            new_worksheet['A4'].value = semaine_text
            
            # Si A4 et B4 sont fusionnées, s'assurer que B4 est vide
            if 'B4' in new_worksheet:
                new_worksheet['B4'].value = None
            
            logger.info(f"Numéro de semaine forcé en A4 (sommier): {semaine_text}")
            print(f"DEBUG SOMMIER: A4 écrit avec succès: {semaine_text}")
        except Exception as e:
            logger.warning(f"Erreur lors de l'ajout du numéro de semaine en A4 (sommier): {e}")
            print(f"DEBUG SOMMIER: Erreur A4 principale: {e}")
            # Tentative alternative : écriture directe sans vérification
            try:
                new_worksheet.cell(row=4, column=1, value=f"sem. {numero_semaine}")
                logger.info(f"Numéro de semaine écrit via cell() en A4 (sommier): sem. {numero_semaine}")
                print(f"DEBUG SOMMIER: A4 écrit via fallback: sem. {numero_semaine}")
            except Exception as e2:
                logger.error(f"Échec total de l'écriture en A4 (sommier): {e2}")
                print(f"DEBUG SOMMIER: Échec total A4: {e2}")
        
        return new_workbook
    
    def save_workbook(self, workbook: openpyxl.Workbook, semaine: str, id_fichier: str) -> str:
        """
        Sauvegarde le workbook sommier avec le nom correct
        
        Args:
            workbook: Workbook à sauvegarder
            semaine: Code semaine
            id_fichier: ID du fichier
            
        Returns:
            Chemin du fichier sauvegardé
        """
        filename = self.generate_filename(semaine, id_fichier)
        
        # Utiliser le répertoire de sortie configuré
        try:
            from config import config
            output_dir = config.get_excel_output_directory()
        except Exception as e:
            logger.warning(f"Impossible d'importer config: {e}. Utilisation du répertoire par défaut.")
            output_dir = "output"
        
        filepath = os.path.join(output_dir, filename)
        
        # Crée le dossier de sortie s'il n'existe pas
        os.makedirs(output_dir, exist_ok=True)
        
        # Auto-ajuster les largeurs de colonnes
        if workbook.active:
            ws = workbook.active
            # Colonnes gauches (coches "X") : largeur compacte basée sur contenu ligne 2
            left_cols = [b[0] for b in self.column_blocks]
            # Colonnes droites (descriptions) : largeur minimale 25 pour lisibilité
            right_cols = [b[1] for b in self.column_blocks]

            for col_letter in left_cols:
                try:
                    cell_value = ws[f'{col_letter}2'].value
                    if cell_value:
                        content_length = len(str(cell_value))
                        adjusted_width = max(content_length + 1, 3)
                        ws.column_dimensions[col_letter].width = adjusted_width
                except Exception as e:
                    logger.warning(f"Erreur ajustement colonne {col_letter}: {e}")

            for col_letter in right_cols:
                try:
                    # S'assurer que les colonnes de description ont au moins 25 de large
                    current_width = ws.column_dimensions[col_letter].width
                    if current_width < 25:
                        ws.column_dimensions[col_letter].width = 25
                except Exception as e:
                    logger.warning(f"Erreur ajustement colonne {col_letter}: {e}")

            # Activer le retour à la ligne sur toutes les cellules de description
            from openpyxl.styles import Alignment
            for col_letter in right_cols:
                col_idx = openpyxl.utils.column_index_from_string(col_letter)
                for row in range(1, 50):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value and isinstance(cell.value, str) and len(str(cell.value)) > 20:
                        cell.alignment = Alignment(wrap_text=True, vertical='center')
        
        workbook.save(filepath)
        logger.info(f"Fichier sommier sauvegardé: {filepath}")
        
        # Retourner le chemin absolu pour éviter les problèmes de résolution
        return os.path.abspath(filepath)
    
    def _detecter_type_sommier_pour_tri(self, config: Dict) -> int:
        """
        Détermine la priorité de tri d'une configuration sommier.
        Ordre : 1/ Motorisé (TT_TENON), 2/ TPR, 3/ LAF 3 côtés, 4/ LAF 4 côtés,
                5/ Lattes dessous, 6/ Lattes dessus, 99/ Autre
        """
        articles = config.get("articles", [])
        description_combined = config.get("description", "")

        # Chercher la description du sommier dans les articles
        sommier_desc = ""
        for art in articles:
            if isinstance(art, dict):
                desc = art.get("description", "")
                if desc.upper().lstrip().startswith("SOMMIER"):
                    sommier_desc = desc
                    break
        if not sommier_desc:
            sommier_desc = description_combined

        desc_upper = sommier_desc.upper()
        desc_norm = re.sub(r'\s+', ' ', desc_upper)

        # Motorisé télescopique → priorité 1
        has_motorise = "MOTORIS" in desc_norm
        has_telescop = "TELESCOP" in desc_norm
        if has_motorise and has_telescop:
            return 1
        # Motorisé non-télescopique (TT_TPR) → priorité 1 aussi
        if has_motorise:
            return 1

        # TPR (relaxation manuelle)
        if "RELAXATION MANUELLE" in desc_norm or "MANUEL" in desc_norm:
            return 2

        # FIXE → déterminer le sous-type
        # LAF 3 côtés
        if "3 COTE" in desc_norm or "3 CÔTÉ" in desc_norm or "3 COTES" in desc_norm:
            return 3
        # LAF 4 côtés
        if "4 COTE" in desc_norm or "4 CÔTÉ" in desc_norm or "4 COTES" in desc_norm:
            return 4
        # Lattes dessous
        if "LATTES DESSOUS" in desc_norm:
            return 5
        # Lattes dessus
        if "LATTES DESSUS" in desc_norm:
            return 6

        # FIXE sans précision → après les motorisés
        if "FIXE" in desc_norm:
            return 7

        return 99

    def _trier_par_type_sommier(self, configurations: List[Dict]) -> List[Dict]:
        """Trie les configurations par type de sommier selon l'ordre standard."""
        return sorted(configurations, key=lambda c: self._detecter_type_sommier_pour_tri(c))

    def import_configurations(self, configurations: List[Dict], semaine: str, id_fichier: str) -> List[str]:
        """
        Importe une liste de configurations sommier dans Excel
        
        Args:
            configurations: Liste des configurations JSON
            semaine: Code semaine
            id_fichier: ID du fichier
            
        Returns:
            Liste des fichiers créés
        """
        created_files = []
        self.total_written_cases = 0

        # Utiliser le répertoire de sortie configuré
        try:
            from config import config
            output_dir = config.get_excel_output_directory()
        except Exception as e:
            logger.warning(f"Impossible d'importer config: {e}. Utilisation du répertoire par défaut.")
            output_dir = "output"
        
        # Cherche le dernier fichier existant pour cette semaine
        self.current_file_index = 1
        while True:
            filename = self.generate_filename(semaine, id_fichier)
            filepath = os.path.join(output_dir, filename)
            if os.path.exists(filepath):
                self.current_file_index += 1
            else:
                break
        
        # Si on a trouvé des fichiers existants, charge le dernier
        if self.current_file_index > 1:
            self.current_file_index -= 1
            filename = self.generate_filename(semaine, id_fichier)
            filepath = os.path.join(output_dir, filename)
            logger.info(f"Chargement du fichier sommier existant: {filepath}")
            self.current_workbook = openpyxl.load_workbook(filepath)
            self.current_worksheet = self.current_workbook.active
            
            # Compte les cas déjà présents
            self.current_case_count = 0
            for left_col, right_col in self.column_blocks:
                if not self.is_block_empty(self.current_worksheet, left_col, right_col):
                    self.current_case_count += 1
            logger.info(f"Fichier sommier existant contient {self.current_case_count} cas")
        else:
            # Charge le template initial
            filename = self.generate_filename(semaine, id_fichier)
            filepath = os.path.join(output_dir, filename)
            logger.info(f"Création d'un nouveau fichier sommier: {filepath}")
            self.current_workbook = self.load_template()
            self.current_worksheet = self.current_workbook.active
            self.current_case_count = 0
            
            # Ajouter le numéro de semaine en A4 pour le premier fichier
            try:
                numero_semaine = semaine.replace('S', '') if semaine.startswith('S') else semaine
                semaine_text = f"sem. {numero_semaine}"
                self.current_worksheet['A4'].value = semaine_text
                if 'B4' in self.current_worksheet:
                    self.current_worksheet['B4'].value = None
                logger.info(f"Numéro de semaine ajouté au premier fichier sommier en A4: {semaine_text}")
                print(f"DEBUG PREMIER FICHIER SOMMIER: A4 écrit avec succès: {semaine_text}")
            except Exception as e:
                logger.error(f"Erreur lors de l'ajout du numéro de semaine au premier fichier sommier: {e}")
                print(f"DEBUG ERREUR PREMIER FICHIER SOMMIER: {e}")
        
        # Trier les configurations par type de sommier
        configurations = self._trier_par_type_sommier(configurations)

        for i, config in enumerate(configurations):
            logger.info(f"Traitement configuration sommier {i+1}/{len(configurations)}")
            
            # Vérifie si on doit créer un nouveau fichier
            if self.current_case_count >= self.max_cases_per_file:
                # Sauvegarde le fichier actuel
                filepath = self.save_workbook(self.current_workbook, semaine, id_fichier)
                created_files.append(filepath)
                
                # Crée un nouveau fichier
                self.current_workbook = self.create_new_file(semaine, id_fichier)
                self.current_worksheet = self.current_workbook.active
            
            # Trouve le prochain bloc vide
            empty_block = self.find_next_empty_block(self.current_worksheet)
            
            if empty_block is None:
                # Tous les blocs sont pleins, crée un nouveau fichier
                filepath = self.save_workbook(self.current_workbook, semaine, id_fichier)
                created_files.append(filepath)
                
                self.current_workbook = self.create_new_file(semaine, id_fichier)
                self.current_worksheet = self.current_workbook.active
                
                # Retrouve le bloc vide dans le nouveau fichier
                empty_block = self.find_next_empty_block(self.current_worksheet)
            
            left_col, right_col = empty_block
            
            # Écrit la configuration dans le bloc
            try:
                self.write_config_to_block(self.current_worksheet, config, left_col, right_col)
                self.current_case_count += 1
                logger.info(f"Configuration sommier {i+1} écrite dans le bloc {left_col}-{right_col}")
            except Exception as e:
                logger.error(f"Erreur lors de l'écriture du sommier {i+1} dans le bloc {left_col}-{right_col}: {e}")
                import traceback
                logger.error(f"Traceback: {traceback.format_exc()}")
                # Continue avec le sommier suivant même en cas d'erreur
                continue
        
        # Sauvegarde le dernier fichier
        if self.current_workbook:
            filepath = self.save_workbook(self.current_workbook, semaine, id_fichier)
            created_files.append(filepath)
        
        if self.total_written_cases != len(configurations):
            msg = (
                f"Nombre de blocs écrits ({self.total_written_cases}) "
                f"différent du nombre de configurations ({len(configurations)})"
            )
            logger.error(msg)
            raise RuntimeError(msg)
        logger.info(f"Import sommier terminé. {len(created_files)} fichier(s) créé(s)")
        return created_files


def main():
    """
    Fonction principale pour tester l'importateur de sommiers
    """
    # Exemple d'utilisation
    importer = ExcelSommierImporter()
    
    # Exemple de configurations JSON pour sommiers
    sample_configs = [
        {
            "Client_D1": "Client Test 1",
            "Client_D2": "Adresse Test 1",
            "Article_D6": "Sommier à lattes",
            "Quantite_D11": "2",
            "Dimensions_D15": "160x200",
            "Type_Sommier_D20": "SOMMIER À LATTES",
            "Materiau_D25": "BOIS",
            "Hauteur_D30": "8",
            "Prix_D35": "400.00"
        },
        {
            "Client_D1": "Client Test 2",
            "Client_D2": "Adresse Test 2",
            "Article_D6": "Sommier tapissier",
            "Quantite_D11": "1",
            "Dimensions_D15": "140x190",
            "Type_Sommier_D20": "SOMMIER TAPISSIER",
            "Materiau_D25": "TAPISSIER",
            "Hauteur_D30": "12",
            "Prix_D35": "300.00"
        }
    ]
    
    try:
        created_files = importer.import_configurations(sample_configs, "S01", "1234")
        print(f"Fichiers sommiers créés: {created_files}")
    except Exception as e:
        logger.error(f"Erreur lors de l'import sommier: {e}")


if __name__ == "__main__":
    main() 
