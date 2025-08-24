#!/usr/bin/env python3
"""
Gestionnaire de mappings pour les configurations Excel
"""

import json
import os
import sys
import logging
from typing import Dict, List, Optional, Tuple
import openpyxl
from datetime import datetime

logger = logging.getLogger(__name__)

# Import asset_utils pour la gestion des chemins
try:
    from .asset_utils import get_config_path
except ImportError:
    try:
        from asset_utils import get_config_path
    except ImportError:
        # Fallback si asset_utils n'est pas disponible
        def get_config_path(relative_path: str) -> Optional[str]:
            try:
                if hasattr(sys, '_MEIPASS'):
                    base_path = sys._MEIPASS
                else:
                    base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                config_path = os.path.join(base_path, 'config', relative_path)
                if os.path.exists(config_path):
                    return config_path
                fallback_path = os.path.join('config', relative_path)
                if os.path.exists(fallback_path):
                    return fallback_path
                return None
            except Exception:
                return None

class MappingManager:
    """
    Gestionnaire centralisé des mappings entre champs pré-import et cellules Excel
    """
    
    def __init__(self, config_dir: str = "config"):
        """
        Initialise le gestionnaire de mappings
        
        Args:
            config_dir: Répertoire de configuration
        """
        self.config_dir = config_dir
        # Utiliser asset_utils pour les chemins
        self.matelas_mappings_file = get_config_path("mappings_matelas.json")
        self.sommiers_mappings_file = get_config_path("mappings_sommiers.json")
        
        # Fallback si asset_utils ne trouve pas les fichiers
        if not self.matelas_mappings_file:
            self.matelas_mappings_file = os.path.join(config_dir, "mappings_matelas.json")
        if not self.sommiers_mappings_file:
            self.sommiers_mappings_file = os.path.join(config_dir, "mappings_sommiers.json")
        
        # Mappings par défaut pour les matelas
        self.default_matelas_mappings = {
            "Client_D1": "D1",
            "Adresse_D3": "D3",
            "numero_D2": "D2",
            "semaine_D5": "D5",
            "lundi_D6": "D6",
            "vendredi_D7": "D7",
            "Hauteur_D22": "D22",
            "dimension_housse_D23": "D23",
            "longueur_D24": "D24",
            "decoupe_noyau_D25": "D25",
            "jumeaux_C10": "C10",
            "jumeaux_D10": "D10",
            "1piece_C11": "C11",
            "1piece_D11": "D11",
            "dosseret_tete_C8": "C8",
            "poignees_C20": "C20",
            "Surmatelas_C45": "C45",
            "emporte_client_C57": "C57",
            "fourgon_C58": "C58",
            "transporteur_C59": "C59"
        }
        
        # Mappings par défaut pour les sommiers
        self.default_sommiers_mappings = {
            "Client_D1": "D1",
            "Adresse_D3": "D3",
            "numero_D2": "D2",
            "semaine_D5": "D5",
            "lundi_D6": "D6",
            "vendredi_D7": "D7",
            "Type_Sommier_D20": "D20",
            "Materiau_D25": "D25",
            "Hauteur_D30": "D30",
            "Dimensions_D35": "D35",
            "Quantite_D40": "D40",
            "Sommier_DansUnLit_D45": "D45",
            "Sommier_Pieds_D50": "D50",
            "emporte_client_C57": "C57",
            "fourgon_C58": "C58",
            "transporteur_C59": "C59"
        }
        
        # Charger les mappings
        self.matelas_mappings = self.load_mappings("matelas")
        self.sommiers_mappings = self.load_mappings("sommiers")
    
    def load_mappings(self, product_type: str) -> Dict[str, str]:
        """
        Charge les mappings depuis le fichier JSON
        
        Args:
            product_type: "matelas" ou "sommiers"
            
        Returns:
            Dictionnaire des mappings
        """
        file_path = self.matelas_mappings_file if product_type == "matelas" else self.sommiers_mappings_file
        default_mappings = self.default_matelas_mappings if product_type == "matelas" else self.default_sommiers_mappings
        
        try:
            if file_path and os.path.exists(file_path):
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    mappings = data.get("mappings", {})
                    logger.info(f"Mappings {product_type} chargés depuis {file_path}")
                    return mappings
            else:
                logger.info(f"Fichier de mappings {product_type} non trouvé ({file_path}), utilisation des valeurs par défaut")
                return default_mappings.copy()
        except Exception as e:
            logger.error(f"Erreur lors du chargement des mappings {product_type}: {e}")
            return default_mappings.copy()
    
    def save_mappings(self, product_type: str, mappings: Dict[str, str]) -> bool:
        """
        Sauvegarde les mappings dans le fichier JSON
        
        Args:
            product_type: "matelas" ou "sommiers"
            mappings: Dictionnaire des mappings à sauvegarder
            
        Returns:
            True si sauvegarde réussie
        """
        # Déterminer le chemin de sauvegarde
        if hasattr(sys, '_MEIPASS'):
            # Mode PyInstaller - sauvegarder dans le dossier de l'exécutable
            base_dir = os.path.dirname(sys.executable) if hasattr(sys, 'executable') else os.getcwd()
            config_dir = os.path.join(base_dir, 'config')
        else:
            # Mode développement
            config_dir = self.config_dir
        
        file_name = "mappings_matelas.json" if product_type == "matelas" else "mappings_sommiers.json"
        file_path = os.path.join(config_dir, file_name)
        
        try:
            # Créer le répertoire si nécessaire
            os.makedirs(config_dir, exist_ok=True)
            
            data = {
                "mappings": mappings,
                "metadata": {
                    "version": "1.0",
                    "last_modified": datetime.now().isoformat(),
                    "description": f"Mappings pour les {product_type}"
                }
            }
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            
            # Mettre à jour les mappings en mémoire
            if product_type == "matelas":
                self.matelas_mappings = mappings
            else:
                self.sommiers_mappings = mappings
            
            logger.info(f"Mappings {product_type} sauvegardés dans {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Erreur lors de la sauvegarde des mappings {product_type}: {e}")
            return False
    
    def get_cell_for_field(self, field_name: str, product_type: str) -> str:
        """
        Retourne la cellule Excel pour un champ donné
        
        Args:
            field_name: Nom du champ (ex: "Client_D1")
            product_type: "matelas" ou "sommiers"
            
        Returns:
            Adresse de cellule (ex: "C1")
        """
        mappings = self.matelas_mappings if product_type == "matelas" else self.sommiers_mappings
        return mappings.get(field_name, "")
    
    def validate_cell_format(self, cell_address: str) -> bool:
        """
        Valide le format d'une adresse de cellule Excel
        
        Args:
            cell_address: Adresse de cellule (ex: "C1", "D25")
            
        Returns:
            True si format valide
        """
        if not cell_address or not isinstance(cell_address, str):
            return False
        
        # Format attendu : lettre(s) + chiffre(s)
        import re
        pattern = r'^[A-Z]+\d+$'
        return bool(re.match(pattern, cell_address))
    
    def validate_cell_exists(self, cell_address: str, template_path: str) -> bool:
        """
        Valide qu'une cellule existe dans le template Excel
        
        Args:
            cell_address: Adresse de cellule (ex: "C1")
            template_path: Chemin vers le template Excel
            
        Returns:
            True si la cellule existe
        """
        try:
            if not os.path.exists(template_path):
                logger.warning(f"Template non trouvé: {template_path}")
                return False
            
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # Extraire colonne et ligne
            import re
            match = re.match(r'([A-Z]+)(\d+)', cell_address)
            if not match:
                return False
            
            col_letter = match.group(1)
            row_num = int(match.group(2))
            
            # Vérifier que la cellule existe
            exists = (row_num <= ws.max_row and 
                     openpyxl.utils.column_index_from_string(col_letter) <= ws.max_column)
            
            wb.close()
            return exists
            
        except Exception as e:
            logger.error(f"Erreur lors de la validation de la cellule {cell_address}: {e}")
            return False
    
    def get_all_fields(self, product_type: str) -> List[str]:
        """
        Retourne la liste de tous les champs disponibles
        
        Args:
            product_type: "matelas" ou "sommiers"
            
        Returns:
            Liste des noms de champs
        """
        if product_type == "matelas":
            return list(self.default_matelas_mappings.keys())
        else:
            return list(self.default_sommiers_mappings.keys())
    
    def reset_to_defaults(self, product_type: str) -> bool:
        """
        Remet les mappings aux valeurs par défaut
        
        Args:
            product_type: "matelas" ou "sommiers"
            
        Returns:
            True si réinitialisation réussie
        """
        default_mappings = (self.default_matelas_mappings if product_type == "matelas" 
                          else self.default_sommiers_mappings)
        
        return self.save_mappings(product_type, default_mappings.copy())
    
    def get_mappings_summary(self, product_type: str) -> Dict:
        """
        Retourne un résumé des mappings pour l'affichage
        
        Args:
            product_type: "matelas" ou "sommiers"
            
        Returns:
            Dictionnaire avec les informations des mappings
        """
        mappings = self.matelas_mappings if product_type == "matelas" else self.sommiers_mappings
        default_mappings = (self.default_matelas_mappings if product_type == "matelas" 
                          else self.default_sommiers_mappings)
        
        return {
            "total_fields": len(mappings),
            "customized_fields": sum(1 for k, v in mappings.items() 
                                   if default_mappings.get(k) != v),
            "last_modified": self._get_last_modified(product_type)
        }
    
    def _get_last_modified(self, product_type: str) -> str:
        """Récupère la date de dernière modification"""
        file_path = self.matelas_mappings_file if product_type == "matelas" else self.sommiers_mappings_file
        
        if file_path and os.path.exists(file_path):
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    return data.get("metadata", {}).get("last_modified", "Inconnue")
            except:
                return "Inconnue"
        return "Jamais" 