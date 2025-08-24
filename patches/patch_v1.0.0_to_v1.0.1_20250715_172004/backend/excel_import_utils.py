"""
Script d'automatisation Excel pour l'import de configurations de matelas
Auteur: Assistant IA
Date: 2024
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os
import json
from typing import Dict, List, Tuple, Optional
import logging
import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(__file__)))

# Configuration du logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ExcelMatelasImporter:
    """
    Classe pour automatiser l'import de configurations de matelas dans Excel
    """
    
    def __init__(self, template_path: str = "template/template_matelas.xlsx", 
                 alignment_mode: str = "intelligent"):
        """
        Initialise l'importateur Excel
        
        Args:
            template_path: Chemin vers le fichier template
            alignment_mode: Mode d'alignement ("intelligent", "global", "none")
                - "intelligent": Alignement spécifique par type de données
                - "global": Centrage de toutes les cellules
                - "none": Aucun alignement personnalisé
        """
        self.template_path = template_path
        self.alignment_mode = alignment_mode
        self.current_file_index = 1
        self.current_case_count = 0
        self.max_cases_per_file = 10
        
        # Charger le mapping manager pour utiliser les mappings sauvegardés
        try:
            from mapping_manager import MappingManager
            self.mapping_manager = MappingManager()
            logger.info("Mapping manager chargé avec succès")
        except ImportError as e:
            logger.warning(f"Impossible de charger le mapping manager: {e}. Utilisation des mappings par défaut.")
            self.mapping_manager = None
        
        # Définition des blocs de colonnes (cas 1-10)
        # Format: (colonne_gauche, colonne_droite)
        self.column_blocks = [
            ('C', 'D'),  # Cas 1
            ('E', 'F'),  # Cas 2
            ('G', 'H'),  # Cas 3
            ('I', 'J'),  # Cas 4
            ('K', 'L'),  # Cas 5
            # M & N sont verrouillés - on saute
            ('O', 'P'),  # Cas 6
            ('Q', 'R'),  # Cas 7
            ('S', 'T'),  # Cas 8
            ('U', 'V'),  # Cas 9
            ('W', 'X'),  # Cas 10
        ]
        
        # Mapping par défaut (utilisé si mapping_manager non disponible)
        # Format: {clé_json: (colonne, ligne)}
        self.default_json_to_cell_mapping = {
            # Informations client
            'Client_D1': ('D', 1),
            'Adresse_D3': ('D', 3),
            
            # Champs commande et dates
            'numero_D2': ('D', 2),
            'semaine_D5': ('D', 5),
            'lundi_D6': ('D', 6),
            'vendredi_D7': ('D', 7),
            
            # Champs matelas
            'Hauteur_D22': ('D', 22),
            
            # Champs détection
            'dosseret_tete_C8': ('C', 8),
            
            # Champs quantité
            'jumeaux_C10': ('C', 10),
            'jumeaux_D10': ('D', 10),
            '1piece_C11': ('C', 11),
            '1piece_D11': ('D', 11),
            
            # Champs housse et matière
            'HSimple_polyester_C13': ('C', 13),
            'HSimple_polyester_D13': ('D', 13),
            'HSimple_tencel_C14': ('C', 14),
            'HSimple_tencel_D14': ('D', 14),
            'HSimple_autre_C15': ('C', 15),
            'HSimple_autre_D15': ('D', 15),
            'Hmat_polyester_C17': ('C', 17),
            'Hmat_polyester_D17': ('D', 17),
            'Hmat_tencel_C18': ('C', 18),
            'Hmat_tencel_D18': ('D', 18),
            'Hmat_luxe3D_C19': ('C', 19),
            'Hmat_luxe3D_D19': ('D', 19),
            
            # Champs poignées
            'poignees_C20': ('C', 20),
            
            # Champs dimensions
            'dimension_housse_D23': ('D', 23),
            'longueur_D24': ('D', 24),
            'decoupe_noyau_D25': ('D', 25),
            
            # Champs noyau et fermeté
            'LN_Ferme_C28': ('C', 28),
            'LN_Medium_C29': ('C', 29),
            'LM7z_Ferme_C30': ('C', 30),
            'LM7z_Medium_C31': ('C', 31),
            'LM3z_Ferme_C32': ('C', 32),
            'LM3z_Medium_C33': ('C', 33),
            'MV_Ferme_C34': ('C', 34),
            'MV_Medium_C35': ('C', 35),
            'MV_Confort_C36': ('C', 36),
            'MR_Ferme_C37': ('C', 37),
            'MR_Medium_C38': ('C', 38),
            'MR_Confort_C39': ('C', 39),
            'SL43_Ferme_C40': ('C', 40),
            'SL43_Medium_C41': ('C', 41),
            
            # Champs surmatelas
            'Surmatelas_C45': ('C', 45),
            
            # Champs opérations
            'emporte_client_C57': ('C', 57),
            'fourgon_C58': ('C', 58),
            'transporteur_C59': ('C', 59),
        }
        
        self.current_workbook = None
        self.current_worksheet = None
        
        # Règles de centrage intelligent par type de données
        # Format: {clé_json: (alignement_horizontal, alignement_vertical)}
        self.alignment_rules = {
            # En-têtes et informations générales - centrés
            'Client_D1': ('center', 'center'),
            'Adresse_D3': ('center', 'center'),
            'numero_D2': ('center', 'center'),
            'semaine_D5': ('center', 'center'),
            'lundi_D6': ('center', 'center'),
            'vendredi_D7': ('center', 'center'),
            
            # Dimensions et mesures - centrés
            'Hauteur_D22': ('center', 'center'),
            'dimension_housse_D23': ('center', 'center'),
            'longueur_D24': ('center', 'center'),
            'decoupe_noyau_D25': ('center', 'center'),
            
            # Quantités - centrées
            'jumeaux_C10': ('center', 'center'),
            'jumeaux_D10': ('center', 'center'),
            '1piece_C11': ('center', 'center'),
            '1piece_D11': ('center', 'center'),
            
            # Types de housse - centrés
            'HSimple_polyester_C13': ('center', 'center'),
            'HSimple_polyester_D13': ('center', 'center'),
            'HSimple_tencel_C14': ('center', 'center'),
            'HSimple_tencel_D14': ('center', 'center'),
            'HSimple_autre_C15': ('center', 'center'),
            'HSimple_autre_D15': ('center', 'center'),
            'Hmat_polyester_C17': ('center', 'center'),
            'Hmat_polyester_D17': ('center', 'center'),
            'Hmat_tencel_C18': ('center', 'center'),
            'Hmat_tencel_D18': ('center', 'center'),
            'Hmat_luxe3D_C19': ('center', 'center'),
            'Hmat_luxe3D_D19': ('center', 'center'),
            
            # Poignées - centrées
            'poignees_C20': ('center', 'center'),
            
            # Détection dosseret/tête - centrée
            'dosseret_tete_C8': ('center', 'center'),
            
            # Types de noyau - centrés
            'LN_Ferme_C28': ('center', 'center'),
            'LN_Medium_C29': ('center', 'center'),
            'LM7z_Ferme_C30': ('center', 'center'),
            'LM7z_Medium_C31': ('center', 'center'),
            'LM3z_Ferme_C32': ('center', 'center'),
            'LM3z_Medium_C33': ('center', 'center'),
            'MV_Ferme_C34': ('center', 'center'),
            'MV_Medium_C35': ('center', 'center'),
            'MV_Confort_C36': ('center', 'center'),
            'MR_Ferme_C37': ('center', 'center'),
            'MR_Medium_C38': ('center', 'center'),
            'MR_Confort_C39': ('center', 'center'),
            'SL43_Ferme_C40': ('center', 'center'),
            'SL43_Medium_C41': ('center', 'center'),
            
            # Surmatelas - centré
            'Surmatelas_C45': ('center', 'center'),
            
            # Opérations - centrées
            'emporte_client_C57': ('center', 'center'),
            'fourgon_C58': ('center', 'center'),
            'transporteur_C59': ('center', 'center'),
        }
        
    def generate_filename(self, semaine: str, id_client: str) -> str:
        """
        Génère le nom de fichier selon le format spécifié
        
        Args:
            semaine: Code semaine (ex: "S01")
            id_client: ID du client (ex: "1234")
            
        Returns:
            Nom de fichier généré
        """
        return f"Matelas_{semaine}_{id_client}_{self.current_file_index}.xlsx"
    
    def load_template(self) -> openpyxl.Workbook:
        """
        Charge le template Excel
        
        Returns:
            Workbook du template
        """
        try:
            logger.info(f"Chargement du template: {self.template_path}")
            workbook = openpyxl.load_workbook(self.template_path)
            return workbook
        except FileNotFoundError:
            logger.error(f"Template non trouvé: {self.template_path}")
            raise
        except Exception as e:
            logger.error(f"Erreur lors du chargement du template: {e}")
            raise
    
    def is_block_empty(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                      left_col: str, right_col: str) -> bool:
        """
        Vérifie si un bloc de colonnes est vide (cellule clé D1 du bloc)
        
        Args:
            worksheet: Feuille de calcul
            left_col: Colonne gauche du bloc
            right_col: Colonne droite du bloc
            
        Returns:
            True si le bloc est vide, False sinon
        """
        # On vérifie la cellule clé (D1 du bloc, donc la colonne droite ligne 1)
        key_cell = f"{right_col}1"
        
        try:
            cell = worksheet[key_cell]
            cell_value = cell.value
            
            # Gestion des cellules fusionnées
            if hasattr(cell, 'value') and cell_value is not None:
                logger.info(f"Vérification bloc {left_col}-{right_col}: cellule {key_cell} = '{cell_value}'")
                return str(cell_value).strip() == ""
            else:
                logger.info(f"Vérification bloc {left_col}-{right_col}: cellule {key_cell} est vide ou fusionnée")
                return True
                
        except AttributeError as e:
            if "'MergedCell' object attribute 'value' is read-only" in str(e):
                logger.info(f"Vérification bloc {left_col}-{right_col}: cellule {key_cell} est fusionnée (considérée comme vide)")
                return True
            else:
                logger.error(f"Erreur lors de la vérification de {key_cell}: {e}")
                return True
        except Exception as e:
            logger.error(f"Erreur lors de la vérification de {key_cell}: {e}")
            return True
    
    def find_next_empty_block(self, worksheet: openpyxl.worksheet.worksheet.Worksheet) -> Optional[Tuple[str, str]]:
        """
        Trouve le prochain bloc vide selon l'ordre défini
        
        Args:
            worksheet: Feuille de calcul
            
        Returns:
            Tuple (colonne_gauche, colonne_droite) du bloc vide, ou None si tous pleins
        """
        for left_col, right_col in self.column_blocks:
            if self.is_block_empty(worksheet, left_col, right_col):
                logger.info(f"Bloc vide trouvé: {left_col}-{right_col}")
                return (left_col, right_col)
        
        logger.warning("Aucun bloc vide trouvé dans le fichier actuel")
        return None
    
    def map_json_to_cells(self, config_json: Dict, left_col: str, right_col: str) -> Dict[str, str]:
        """
        Mappe les données JSON vers les cellules du bloc spécifié
        
        Args:
            config_json: Configuration JSON du matelas
            left_col: Colonne gauche du bloc
            right_col: Colonne droite du bloc
            
        Returns:
            Dictionnaire {cellule: valeur} pour l'écriture
        """
        cell_mapping = {}
        
        # Utiliser le mapping manager si disponible, sinon le mapping par défaut
        if self.mapping_manager:
            # Récupérer les mappings sauvegardés pour les matelas
            saved_mappings = self.mapping_manager.matelas_mappings
            logger.info(f"Utilisation des mappings sauvegardés: {len(saved_mappings)} champs")
        else:
            saved_mappings = self.default_json_to_cell_mapping
            logger.info("Utilisation des mappings par défaut")

        # Convertir les mappings sauvegardés au format attendu
        json_to_cell_mapping = {}
        for field_name, cell_address in saved_mappings.items():
            if not cell_address:  # Champ ignoré
                continue
                
            # Extraire colonne et ligne de l'adresse de cellule
            import re
            match = re.match(r'([A-Z]+)(\d+)', cell_address)
            if match:
                col_letter = match.group(1)
                row_num = int(match.group(2))
                
                # Déterminer si c'est une colonne C ou D dans le template
                if col_letter in ['C', 'E', 'G', 'I', 'K', 'O', 'Q', 'S', 'U', 'W']:
                    template_col = 'C'
                else:
                    template_col = 'D'
                
                json_to_cell_mapping[field_name] = (template_col, row_num)

        for json_key, (template_col, row) in json_to_cell_mapping.items():
            # On remplace la colonne template par la colonne du bloc
            if template_col == 'D':
                actual_col = right_col
            else:
                actual_col = left_col
                
            cell_address = f"{actual_col}{row}"
            value = config_json.get(json_key, "")
            
            # Ne pas écrire les champs vides (ignorés)
            if value:
                cell_mapping[cell_address] = value
                logger.debug(f"Mapping: {json_key} -> {cell_address} = {value}")
        
        return cell_mapping
    
    def apply_cell_alignment(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                           cell_address: str, json_key: str) -> None:
        """
        Applique l'alignement intelligent à une cellule selon les règles définies
        
        Args:
            worksheet: Feuille de calcul
            cell_address: Adresse de la cellule (ex: "D1")
            json_key: Clé JSON correspondante (ex: "Client_D1")
        """
        try:
            # Vérifie si une règle d'alignement existe pour cette clé
            if json_key in self.alignment_rules:
                horizontal, vertical = self.alignment_rules[json_key]
                
                # Applique l'alignement
                cell = worksheet[cell_address]
                cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)
                
                logger.debug(f"Alignement appliqué à {cell_address}: {horizontal}/{vertical}")
            else:
                # Alignement par défaut pour les cellules sans règle spécifique
                cell = worksheet[cell_address]
                cell.alignment = Alignment(horizontal='center', vertical='center')
                logger.debug(f"Alignement par défaut appliqué à {cell_address}")
                
        except Exception as e:
            logger.warning(f"Erreur lors de l'application de l'alignement à {cell_address}: {e}")
    
    def center_block_cells(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                          left_col: str, right_col: str) -> None:
        """
        Centre toutes les cellules d'un bloc (option globale)
        
        Args:
            worksheet: Feuille de calcul
            left_col: Colonne gauche du bloc
            right_col: Colonne droite du bloc
        """
        try:
            # Définit la plage de lignes à traiter (basée sur le mapping)
            rows_to_process = set()
            if self.mapping_manager:
                saved_mappings = self.mapping_manager.matelas_mappings
            else:
                saved_mappings = self.default_json_to_cell_mapping

            # Extraire les lignes des mappings
            import re
            for field_name, cell_address in saved_mappings.items():
                if cell_address:  # Ignorer les champs vides
                    match = re.match(r'([A-Z]+)(\d+)', cell_address)
                    if match:
                        row_num = int(match.group(2))
                        rows_to_process.add(row_num)
            
            # Applique le centrage à toutes les cellules du bloc
            for row in rows_to_process:
                for col in [left_col, right_col]:
                    cell_address = f"{col}{row}"
                    try:
                        cell = worksheet[cell_address]
                        if hasattr(cell, 'alignment'):
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            logger.debug(f"Centrage global appliqué à {cell_address}")
                    except Exception as e:
                        logger.debug(f"Impossible de centrer {cell_address}: {e}")
                        
        except Exception as e:
            logger.warning(f"Erreur lors du centrage global du bloc {left_col}-{right_col}: {e}")
    
    def write_config_to_block(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                             config_json: Dict, left_col: str, right_col: str) -> None:
        """
        Écrit une configuration dans un bloc spécifique
        
        Args:
            worksheet: Feuille de calcul
            config_json: Configuration JSON du matelas
            left_col: Colonne gauche du bloc
            right_col: Colonne droite du bloc
        """
        cell_mapping = self.map_json_to_cells(config_json, left_col, right_col)
        
        for cell_address, value in cell_mapping.items():
            try:
                # Vérifie si la cellule est fusionnée
                cell = worksheet[cell_address]
                
                # Vérifie si c'est une cellule fusionnée
                if hasattr(cell, 'value'):
                    # Cellule normale, on peut écrire
                    worksheet[cell_address] = value
                    
                    # Applique l'alignement intelligent
                    # Trouve la clé JSON correspondante à cette cellule
                    json_key = None
                    if self.mapping_manager:
                        saved_mappings = self.mapping_manager.matelas_mappings
                    else:
                        saved_mappings = self.default_json_to_cell_mapping

                    # Créer un mapping inverse pour trouver la clé JSON
                    import re
                    for field_name, mapping_cell_address in saved_mappings.items():
                        if mapping_cell_address:  # Ignorer les champs vides
                            match = re.match(r'([A-Z]+)(\d+)', mapping_cell_address)
                            if match:
                                col_letter = match.group(1)
                                row_num = int(match.group(2))
                                
                                # Déterminer la colonne template
                                if col_letter in ['C', 'E', 'G', 'I', 'K', 'O', 'Q', 'S', 'U', 'W']:
                                    template_col = 'C'
                                else:
                                    template_col = 'D'
                                
                                # Calculer l'adresse réelle dans le bloc
                                if template_col == 'D':
                                    actual_col = right_col
                                else:
                                    actual_col = left_col
                                
                                if f"{actual_col}{row_num}" == cell_address:
                                    json_key = field_name
                                    break
                    
                    # Applique l'alignement selon le mode choisi
                    if self.alignment_mode == "intelligent":
                        if json_key:
                            self.apply_cell_alignment(worksheet, cell_address, json_key)
                    elif self.alignment_mode == "global":
                        # Le centrage global sera appliqué après l'écriture de toutes les cellules
                        pass
                    
                    logger.info(f"Écriture: {cell_address} = {value}")
                else:
                    logger.warning(f"Cellule {cell_address} non accessible (fusionnée ou protégée)")
                    
            except AttributeError as e:
                if "'MergedCell' object attribute 'value' is read-only" in str(e):
                    logger.warning(f"Cellule {cell_address} est fusionnée, impossible d'écrire")
                else:
                    logger.error(f"Erreur lors de l'écriture de {cell_address}: {e}")
            except Exception as e:
                logger.error(f"Erreur lors de l'écriture de {cell_address}: {e}")
        
        # Applique le centrage global si le mode est activé
        if self.alignment_mode == "global":
            self.center_block_cells(worksheet, left_col, right_col)
            logger.info(f"Centrage global appliqué au bloc {left_col}-{right_col}")
    
    def create_new_file(self, semaine: str, id_fichier: str) -> openpyxl.Workbook:
        """
        Crée un nouveau fichier Excel à partir du template
        
        Args:
            semaine: Code semaine
            id_fichier: ID du fichier
            
        Returns:
            Nouveau workbook
        """
        self.current_file_index += 1
        self.current_case_count = 0
        
        logger.info(f"Création nouveau fichier: index {self.current_file_index}")
        
        # Charge le template pour le nouveau fichier
        new_workbook = self.load_template()
        
        return new_workbook
    
    def save_workbook(self, workbook: openpyxl.Workbook, semaine: str, id_fichier: str) -> str:
        """
        Sauvegarde le workbook avec le nom correct
        
        Args:
            workbook: Workbook à sauvegarder
            semaine: Code semaine
            id_fichier: ID du fichier
            
        Returns:
            Chemin du fichier sauvegardé
        """
        filename = self.generate_filename(semaine, id_fichier)
        
        # Utiliser le répertoire de sortie configuré
        try:
            from config import config
            output_dir = config.get_excel_output_directory()
        except Exception as e:
            logger.warning(f"Impossible d'importer config: {e}. Utilisation du répertoire par défaut.")
            output_dir = "output"
        
        filepath = os.path.join(output_dir, filename)
        
        # Crée le dossier de sortie s'il n'existe pas
        os.makedirs(output_dir, exist_ok=True)
        
        workbook.save(filepath)
        logger.info(f"Fichier sauvegardé: {filepath}")
        
        return filepath
    
    def import_configurations(self, configurations: List[Dict], semaine: str, id_fichier: str) -> List[str]:
        """
        Importe une liste de configurations dans Excel
        
        Args:
            configurations: Liste des configurations JSON
            semaine: Code semaine
            id_fichier: ID du fichier (pas du client)
            
        Returns:
            Liste des fichiers créés
        """
        created_files = []

        # Import de config avec gestion d'erreur
        try:
            from config import config
        except Exception as e:
            logger.warning(f"Impossible d'importer config: {e}. Aucun ordre de noyaux appliqué.")
            # Créer un objet config par défaut
            class DefaultConfig:
                def get_noyau_order(self):
                    return []
            config = DefaultConfig()

        # Vérification explicite de l'objet config
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Type de config avant tri noyaux: {type(config)}")
        if not hasattr(config, 'get_noyau_order'):
            raise RuntimeError(f"L'objet config n'a pas la méthode get_noyau_order. Type: {type(config)}")

        # Note: Le tri global est maintenant fait au niveau de backend_interface.py
        # Les configurations arrivent déjà triées ici
        logger.info(f"Nombre de configurations reçues (déjà triées): {len(configurations)}")
        
        # Afficher les noyaux présents dans les configurations pour debug
        noyaux_presents = []
        for conf in configurations:
            noyau = conf.get('Noyau') or conf.get('Type') or conf.get('noyau')
            if noyau:
                noyaux_presents.append(noyau)
        logger.info(f"Noyaux présents dans les configurations: {noyaux_presents}")
        
        # Utiliser le répertoire de sortie configuré
        try:
            from config import config
            output_dir = config.get_excel_output_directory()
        except Exception as e:
            logger.warning(f"Impossible d'importer config: {e}. Utilisation du répertoire par défaut.")
            output_dir = "output"
        
        # Cherche le dernier fichier existant pour cette semaine
        self.current_file_index = 1
        while True:
            filename = self.generate_filename(semaine, id_fichier)
            filepath = os.path.join(output_dir, filename)
            if os.path.exists(filepath):
                self.current_file_index += 1
            else:
                break
        
        # Si on a trouvé des fichiers existants, charge le dernier
        if self.current_file_index > 1:
            self.current_file_index -= 1
            filename = self.generate_filename(semaine, id_fichier)
            filepath = os.path.join(output_dir, filename)
            logger.info(f"Chargement du fichier existant: {filepath}")
            self.current_workbook = openpyxl.load_workbook(filepath)
            self.current_worksheet = self.current_workbook.active
            
            # Compte les cas déjà présents
            self.current_case_count = 0
            for left_col, right_col in self.column_blocks:
                if not self.is_block_empty(self.current_worksheet, left_col, right_col):
                    self.current_case_count += 1
            logger.info(f"Fichier existant contient {self.current_case_count} cas")
        else:
            # Charge le template initial
            filename = self.generate_filename(semaine, id_fichier)
            filepath = os.path.join(output_dir, filename)
            logger.info(f"Création d'un nouveau fichier: {filepath}")
            self.current_workbook = self.load_template()
            self.current_worksheet = self.current_workbook.active
            self.current_case_count = 0
        
        for i, config in enumerate(configurations):
            logger.info(f"Traitement configuration {i+1}/{len(configurations)}")
            
            # Vérifie si on doit créer un nouveau fichier
            if self.current_case_count >= self.max_cases_per_file:
                # Sauvegarde le fichier actuel
                filepath = self.save_workbook(self.current_workbook, semaine, id_fichier)
                created_files.append(filepath)
                
                # Crée un nouveau fichier
                self.current_workbook = self.create_new_file(semaine, id_fichier)
                self.current_worksheet = self.current_workbook.active
            
            # Trouve le prochain bloc vide
            empty_block = self.find_next_empty_block(self.current_worksheet)
            
            if empty_block is None:
                # Tous les blocs sont pleins, crée un nouveau fichier
                filepath = self.save_workbook(self.current_workbook, semaine, id_fichier)
                created_files.append(filepath)
                
                self.current_workbook = self.create_new_file(semaine, id_fichier)
                self.current_worksheet = self.current_workbook.active
                
                # Retrouve le bloc vide dans le nouveau fichier
                empty_block = self.find_next_empty_block(self.current_worksheet)
            
            left_col, right_col = empty_block
            
            # Écrit la configuration dans le bloc
            self.write_config_to_block(self.current_worksheet, config, left_col, right_col)
            
            self.current_case_count += 1
            logger.info(f"Configuration {i+1} écrite dans le bloc {left_col}-{right_col}")
        
        # Sauvegarde le dernier fichier
        if self.current_workbook:
            filepath = self.save_workbook(self.current_workbook, semaine, id_fichier)
            created_files.append(filepath)
        
        logger.info(f"Import terminé. {len(created_files)} fichier(s) créé(s)")
        return created_files


def main():
    """
    Fonction principale pour tester l'importateur
    """
    # Exemple d'utilisation
    importer = ExcelMatelasImporter()
    
    # Exemple de configurations JSON
    sample_configs = [
        {
            "Client_D1": "Client Test 1",
            "Client_D2": "Adresse Test 1",
            "Article_D6": "Matelas Premium",
            "Quantite_D11": "2",
            "Dimensions_D15": "160x200",
            "Prix_D26": "1500.00"
        },
        {
            "Client_D1": "Client Test 2",
            "Client_D2": "Adresse Test 2",
            "Article_D6": "Matelas Standard",
            "Quantite_D11": "1",
            "Dimensions_D15": "140x190",
            "Prix_D26": "800.00"
        }
    ]
    
    try:
        created_files = importer.import_configurations(sample_configs, "S01", "1234")
        print(f"Fichiers créés: {created_files}")
    except Exception as e:
        logger.error(f"Erreur lors de l'import: {e}")


if __name__ == "__main__":
    main() 