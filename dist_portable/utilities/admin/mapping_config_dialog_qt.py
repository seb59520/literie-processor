#!/usr/bin/env python3
"""
Interface de configuration des mappings Excel avec PyQt6
"""

import sys
import os
import json
from typing import Dict, List
import openpyxl
from datetime import datetime

from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QGridLayout, QLabel, QLineEdit,
    QPushButton, QTabWidget, QWidget, QTableWidget, QTableWidgetItem,
    QHeaderView, QMessageBox, QGroupBox, QRadioButton, QButtonGroup,
    QScrollArea, QFrame, QSplitter, QTextEdit, QComboBox
)
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QFont, QPixmap, QPainter, QColor, QPen

# Ajouter le répertoire backend au path
backend_dir = os.path.join(os.path.dirname(__file__), "..", "..", "backend")
if backend_dir not in sys.path:
    sys.path.insert(0, backend_dir)

try:
    from mapping_manager import MappingManager
except ImportError:
    # Fallback pour les imports relatifs
    try:
        from backend.mapping_manager import MappingManager
    except ImportError:
        # Fallback pour les imports absolus
        sys.path.append(os.path.join(os.path.dirname(__file__), "..", ".."))
        from backend.mapping_manager import MappingManager

class ExcelPreviewWidget(QWidget):
    """Widget de prévisualisation Excel personnalisé avec zoom"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.worksheet = None
        self.mappings = {}
        self.cell_size = 80
        self.row_height = 25
        self.start_x = 50
        self.start_y = 50
        self.setMinimumSize(800, 600)
        self.zoom_factor = 1.0
    
    def set_data(self, worksheet, mappings):
        self.worksheet = worksheet
        self.mappings = mappings
        self.update()
    
    def set_zoom(self, factor):
        self.zoom_factor = max(0.5, min(3.0, factor))
        self.update()
    
    def paintEvent(self, event):
        if not self.worksheet:
            return
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        # Appliquer le zoom
        cell_size = int(self.cell_size * self.zoom_factor)
        row_height = int(self.row_height * self.zoom_factor)
        start_x = int(self.start_x * self.zoom_factor)
        start_y = int(self.start_y * self.zoom_factor)
        # Couleurs
        header_color = QColor(224, 224, 224)  # #E0E0E0
        mapped_color = QColor(144, 238, 144)  # #90EE90
        default_color = QColor(255, 255, 255)  # #FFFFFF
        border_color = QColor(128, 128, 128)   # #808080
        # En-têtes de colonnes
        for col in range(1, min(10, self.worksheet.max_column + 1)):
            col_letter = openpyxl.utils.get_column_letter(col)
            x = start_x + (col - 1) * cell_size
            
            # En-tête de colonne
            painter.fillRect(x, start_y, cell_size, row_height, header_color)
            painter.setPen(QPen(Qt.GlobalColor.black, 1))
            painter.drawRect(x, start_y, cell_size, row_height)
            
            # Texte de l'en-tête
            painter.drawText(x, start_y, cell_size, row_height,
                           Qt.AlignmentFlag.AlignCenter, col_letter)
        
        # Dessiner les lignes et cellules
        for row in range(1, min(60, self.worksheet.max_row + 1)):
            y = start_y + row * row_height
            
            # En-tête de ligne
            painter.fillRect(int(start_x - 30 * self.zoom_factor), int(y), int(30 * self.zoom_factor), int(row_height), header_color)
            painter.drawRect(int(start_x - 30 * self.zoom_factor), int(y), int(30 * self.zoom_factor), int(row_height))
            painter.drawText(int(start_x - 30 * self.zoom_factor), int(y), int(30 * self.zoom_factor), int(row_height),
                           Qt.AlignmentFlag.AlignCenter, str(row))
            
            # Cellules
            for col in range(1, min(10, self.worksheet.max_column + 1)):
                col_letter = openpyxl.utils.get_column_letter(col)
                x = start_x + (col - 1) * cell_size
                cell_address = f"{col_letter}{row}"
                
                # Déterminer la couleur
                is_mapped = cell_address in self.mappings.values()
                fill_color = mapped_color if is_mapped else default_color
                
                # Dessiner la cellule
                painter.fillRect(int(x), int(y), int(cell_size), int(row_height), fill_color)
                painter.setPen(QPen(border_color, 1))
                painter.drawRect(int(x), int(y), int(cell_size), int(row_height))
                
                # Afficher le contenu si c'est une cellule mappée
                if is_mapped:
                    # Trouver le champ correspondant
                    field_name = None
                    for field, cell in self.mappings.items():
                        if cell == cell_address:
                            field_name = field
                            break
                    
                    if field_name:
                        # Afficher le nom du champ (tronqué)
                        display_text = field_name.replace("_", "\n")
                        painter.setPen(QPen(Qt.GlobalColor.black, 1))
                        painter.setFont(QFont("Arial", int(8 * self.zoom_factor)))
                        painter.drawText(int(x), int(y), int(cell_size), int(row_height),
                                       Qt.AlignmentFlag.AlignCenter, display_text)
        
        # Légende
        legend_y = start_y + 21 * row_height + 20 * self.zoom_factor
        painter.fillRect(int(start_x), int(legend_y), int(200 * self.zoom_factor), int(60 * self.zoom_factor), default_color)
        painter.setPen(QPen(Qt.GlobalColor.black, 1))
        painter.drawRect(int(start_x), int(legend_y), int(200 * self.zoom_factor), int(60 * self.zoom_factor))
        
        painter.setFont(QFont("Arial", int(10 * self.zoom_factor), QFont.Weight.Bold))
        painter.drawText(int(start_x + 10 * self.zoom_factor), int(legend_y + 10 * self.zoom_factor), "Légende:")
        
        painter.fillRect(int(start_x + 10 * self.zoom_factor), int(legend_y + 25 * self.zoom_factor), int(20 * self.zoom_factor), int(20 * self.zoom_factor), mapped_color)
        painter.drawRect(int(start_x + 10 * self.zoom_factor), int(legend_y + 25 * self.zoom_factor), int(20 * self.zoom_factor), int(20 * self.zoom_factor))
        painter.setFont(QFont("Arial", int(9 * self.zoom_factor)))
        painter.drawText(int(start_x + 35 * self.zoom_factor), int(legend_y + 35 * self.zoom_factor), "Champ mappé")

class MappingConfigDialog(QDialog):
    """
    Interface de configuration des mappings avec PyQt6
    """
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.mapping_manager = MappingManager()
        
        # Variables
        self.current_product_type = "matelas"
        self.matelas_mappings = self.mapping_manager.matelas_mappings.copy()
        self.sommiers_mappings = self.mapping_manager.sommiers_mappings.copy()
        
        # Templates
        self.matelas_template = "template/template_matelas.xlsx"
        self.sommiers_template = "template/template_sommier.xlsx"
        
        self.setup_ui()
        self.load_preview()
    
    def setup_ui(self):
        """Nouvelle interface utilisateur compacte et moderne"""
        self.setWindowTitle("Configuration des Mappings Excel")
        self.setModal(True)
        self.resize(1200, 800)

        # Layout principal vertical
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(20, 12, 20, 16)
        main_layout.setSpacing(0)

        # --- EN-TÊTE ---
        header_layout = QVBoxLayout()
        header_layout.setSpacing(2)
        title_label = QLabel("Configuration des Mappings Excel")
        title_label.setFont(QFont("Arial", 18, QFont.Weight.Bold))
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        subtitle_label = QLabel("Associez les champs du pré-import aux cellules Excel")
        subtitle_label.setFont(QFont("Arial", 11))
        subtitle_label.setStyleSheet("color: #666;")
        subtitle_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        header_layout.addWidget(title_label)
        header_layout.addWidget(subtitle_label)
        main_layout.addLayout(header_layout)
        main_layout.addSpacing(8)

        # --- ZONE CENTRALE ---
        center_layout = QHBoxLayout()
        center_layout.setSpacing(16)

        # --- PANEL GAUCHE ---
        left_panel = QFrame()
        left_panel.setFrameShape(QFrame.Shape.StyledPanel)
        left_panel.setMinimumWidth(340)
        left_layout = QVBoxLayout(left_panel)
        left_layout.setSpacing(10)
        left_layout.setContentsMargins(8, 8, 8, 8)

        # Sélecteur de type de produit
        product_group = QGroupBox("Type de produit")
        product_layout = QVBoxLayout(product_group)
        self.matelas_radio = QRadioButton("Matelas")
        self.sommiers_radio = QRadioButton("Sommiers")
        self.matelas_radio.setChecked(True)
        self.product_group = QButtonGroup()
        self.product_group.addButton(self.matelas_radio)
        self.product_group.addButton(self.sommiers_radio)
        self.product_group.buttonClicked.connect(self.on_product_type_changed)
        product_layout.addWidget(self.matelas_radio)
        product_layout.addWidget(self.sommiers_radio)
        left_layout.addWidget(product_group)

        # Informations
        self.info_group = QGroupBox("Informations")
        self.info_layout = QVBoxLayout(self.info_group)
        self.info_label = QLabel("")
        self.info_layout.addWidget(self.info_label)
        left_layout.addWidget(self.info_group)

        # Tableau des mappings
        mappings_group = QGroupBox("Mappings")
        mappings_layout = QVBoxLayout(mappings_group)
        from PyQt6.QtWidgets import QPushButton
        self.mappings_table = QTableWidget()
        self.mappings_table.setColumnCount(4)
        self.mappings_table.setHorizontalHeaderLabels(["Champ pré-import", "Cellule Excel", "Statut", "Action"])
        self.mappings_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.mappings_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.mappings_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.mappings_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        mappings_layout.addWidget(self.mappings_table)
        left_layout.addWidget(mappings_group, stretch=1)

        center_layout.addWidget(left_panel, stretch=0)

        # --- PANEL DROIT ---
        right_panel = QFrame()
        right_panel.setFrameShape(QFrame.Shape.StyledPanel)
        right_layout = QVBoxLayout(right_panel)
        right_layout.setSpacing(8)
        right_layout.setContentsMargins(8, 8, 8, 8)

        # Titre prévisualisation
        self.preview_title = QLabel("")
        self.preview_title.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        self.preview_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        right_layout.addWidget(self.preview_title)

        # Zoom + Légende sur la même ligne
        zoom_layout = QHBoxLayout()
        self.zoom_out_btn = QPushButton("-")
        self.zoom_out_btn.setFixedWidth(30)
        self.zoom_out_btn.setStyleSheet("background: white; color: black; border: 1px solid #bbb; font-size: 18px;")
        self.zoom_out_btn.clicked.connect(lambda: self.change_zoom(-0.1))
        self.zoom_in_btn = QPushButton("+")
        self.zoom_in_btn.setFixedWidth(30)
        self.zoom_in_btn.setStyleSheet("background: white; color: black; border: 1px solid #bbb; font-size: 18px;")
        self.zoom_in_btn.clicked.connect(lambda: self.change_zoom(0.1))
        self.zoom_label = QLabel("100%")
        self.zoom_label.setFixedWidth(50)
        zoom_layout.addWidget(self.zoom_out_btn)
        zoom_layout.addWidget(self.zoom_label)
        zoom_layout.addWidget(self.zoom_in_btn)
        # Légende à droite
        self.legend_label = QLabel("<span style='background:#90EE90;padding:2px 8px;border-radius:4px;'>Champ mappé</span>")
        self.legend_label.setFont(QFont("Arial", 10))
        self.legend_label.setStyleSheet("color: #444; margin-left: 16px;")
        zoom_layout.addSpacing(16)
        zoom_layout.addWidget(self.legend_label)
        zoom_layout.addStretch()
        right_layout.addLayout(zoom_layout)
        # Widget de prévisualisation Excel (scrollable, hauteur augmentée)
        self.preview_widget = ExcelPreviewWidget()
        self.preview_widget.setMinimumHeight(1600)  # Pour scroller jusqu'à la ligne 60
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setWidget(self.preview_widget)
        right_layout.addWidget(self.scroll_area, stretch=1)

        center_layout.addWidget(right_panel, stretch=1)
        main_layout.addLayout(center_layout, stretch=1)
        main_layout.addSpacing(8)

        # --- PIED DE PAGE ---
        footer_layout = QHBoxLayout()
        self.save_button = QPushButton("Sauvegarder")
        self.save_button.setDefault(True)
        self.save_button.setStyleSheet("margin-right: 16px;")
        self.save_button.clicked.connect(self.save_mappings)
        self.reset_button = QPushButton("Restaurer défauts")
        self.reset_button.clicked.connect(self.reset_to_defaults)
        self.close_button = QPushButton("Fermer")
        self.close_button.clicked.connect(self.reject)
        footer_layout.addWidget(self.save_button)
        footer_layout.addWidget(self.reset_button)
        footer_layout.addStretch()
        footer_layout.addWidget(self.close_button)
        main_layout.addLayout(footer_layout)

        # Initialiser l'affichage
        self.on_product_type_changed()
    
    def on_product_type_changed(self):
        """Appelé quand le type de produit change"""
        if self.matelas_radio.isChecked():
            self.current_product_type = "matelas"
        else:
            self.current_product_type = "sommiers"
        
        self.load_mappings_table()
        self.load_preview()
        self.update_info()
    
    def load_mappings_table(self):
        """Charge le tableau des mappings"""
        # Obtenir les mappings actuels
        mappings = self.matelas_mappings if self.current_product_type == "matelas" else self.sommiers_mappings
        fields = self.mapping_manager.get_all_fields(self.current_product_type)
        
        # Configurer le tableau
        self.mappings_table.setRowCount(len(fields))
        
        # Remplir le tableau
        for i, field in enumerate(fields):
            # Champ pré-import
            field_item = QTableWidgetItem(field)
            field_item.setFlags(field_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.mappings_table.setItem(i, 0, field_item)
            
            # Cellule Excel
            cell_item = QTableWidgetItem(mappings.get(field, ""))
            self.mappings_table.setItem(i, 1, cell_item)
            
            # Statut
            status_item = QTableWidgetItem("")
            self.mappings_table.setItem(i, 2, status_item)

            # Action : bouton Ignorer
            ignore_btn = QPushButton("Ignorer")
            ignore_btn.setStyleSheet("background: #eee; color: #333; border: 1px solid #bbb; border-radius: 6px; padding: 2px 8px;")
            def make_ignore_func(row=i):
                def ignore():
                    self.mappings_table.item(row, 1).setText("")
                    self.validate_field(field)
                    self.load_preview()
                return ignore
            ignore_btn.clicked.connect(make_ignore_func(i))
            self.mappings_table.setCellWidget(i, 3, ignore_btn)
        
        # Connecter le signal de modification
        self.mappings_table.itemChanged.connect(self.on_cell_changed)
        
        # Valider tous les champs
        self.validate_all_fields()
    
    def on_cell_changed(self, item):
        """Appelé quand une cellule change"""
        if item.column() == 1:  # Colonne cellule Excel
            row = item.row()
            field_item = self.mappings_table.item(row, 0)
            if field_item:
                field_name = field_item.text()
                self.validate_field(field_name)
                self.load_preview()
    
    def validate_field(self, field_name):
        """Valide un champ spécifique"""
        # Trouver la ligne du champ
        for row in range(self.mappings_table.rowCount()):
            field_item = self.mappings_table.item(row, 0)
            if field_item and field_item.text() == field_name:
                cell_item = self.mappings_table.item(row, 1)
                status_item = self.mappings_table.item(row, 2)
                
                if cell_item and status_item:
                    cell_address = cell_item.text().strip().upper()
                    
                    # Champ ignoré
                    if not cell_address:
                        status_item.setText("Ignoré")
                        status_item.setForeground(QColor(128, 128, 128))
                        return True
                    # Validation du format
                    if not self.mapping_manager.validate_cell_format(cell_address):
                        status_item.setText("❌ Format invalide")
                        status_item.setForeground(QColor(255, 0, 0))
                        return False
                    
                    # Validation de l'existence dans le template
                    template_path = self.matelas_template if self.current_product_type == "matelas" else self.sommiers_template
                    
                    if not self.mapping_manager.validate_cell_exists(cell_address, template_path):
                        status_item.setText("⚠️ Cellule inexistante")
                        status_item.setForeground(QColor(255, 165, 0))
                        return False
                    
                    status_item.setText("✅ Valide")
                    status_item.setForeground(QColor(0, 128, 0))
                    return True
                
                break
        
        return False
    
    def validate_all_fields(self):
        """Valide tous les champs"""
        all_valid = True
        for row in range(self.mappings_table.rowCount()):
            field_item = self.mappings_table.item(row, 0)
            cell_item = self.mappings_table.item(row, 1)
            status_item = self.mappings_table.item(row, 2)
            if field_item and cell_item and status_item:
                cell_address = cell_item.text().strip().upper()
                if not cell_address:
                    status_item.setText("Ignoré")
                    status_item.setForeground(QColor("#888"))
                elif not self.mapping_manager.validate_cell_format(cell_address):
                    status_item.setText("❌ Format invalide")
                    status_item.setForeground(QColor("red"))
                    all_valid = False
                elif not self.mapping_manager.validate_cell_exists(cell_address, self.matelas_template if self.current_product_type == "matelas" else self.sommiers_template):
                    status_item.setText("⚠️ Cellule inexistante")
                    status_item.setForeground(QColor(255, 165, 0))
                    all_valid = False
                else:
                    status_item.setText("✅ Valide")
                    status_item.setForeground(QColor("green"))
        return all_valid
    
    def update_info(self):
        """Met à jour les informations"""
        summary = self.mapping_manager.get_mappings_summary(self.current_product_type)
        
        info_text = f"Total: {summary['total_fields']} champs\n"
        info_text += f"Personnalisés: {summary['customized_fields']}\n"
        info_text += f"Dernière modif: {summary['last_modified']}"
        
        self.info_label.setText(info_text)
    
    def load_preview(self):
        """Charge la prévisualisation Excel"""
        product_type = self.current_product_type
        template_path = self.matelas_template if product_type == "matelas" else self.sommiers_template
        
        # Mettre à jour le titre
        template_name = "Matelas" if product_type == "matelas" else "Sommiers"
        self.preview_title.setText(f"Prévisualisation - Template {template_name}")
        
        try:
            if not os.path.exists(template_path):
                return
            
            # Charger le template
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # Obtenir les mappings actuels depuis le tableau
            mappings = {}
            for row in range(self.mappings_table.rowCount()):
                field_item = self.mappings_table.item(row, 0)
                cell_item = self.mappings_table.item(row, 1)
                
                if field_item and cell_item:
                    field_name = field_item.text()
                    cell_address = cell_item.text().strip().upper()
                    if cell_address:
                        mappings[field_name] = cell_address
            
            # Mettre à jour la prévisualisation
            self.preview_widget.set_data(ws, mappings)
            
            wb.close()
            
        except Exception as e:
            print(f"Erreur de prévisualisation: {e}")
    
    def save_mappings(self):
        """Sauvegarde les mappings"""
        product_type = self.current_product_type
        
        # Récupérer les mappings depuis le tableau
        mappings = {}
        for row in range(self.mappings_table.rowCount()):
            field_item = self.mappings_table.item(row, 0)
            cell_item = self.mappings_table.item(row, 1)
            
            if field_item and cell_item:
                field_name = field_item.text()
                cell_address = cell_item.text().strip().upper()
                # On n'ajoute que les champs mappés (cellule non vide)
                if cell_address:
                    mappings[field_name] = cell_address
        
        all_valid = True
        # Valider tous les champs
        for row in range(self.mappings_table.rowCount()):
            field_item = self.mappings_table.item(row, 0)
            if field_item:
                # On considère qu'un champ ignoré (cellule vide) est valide
                if not self.validate_field(field_item.text()):
                    all_valid = False
        
        if not all_valid:
            QMessageBox.warning(self, "Validation", "Certains champs ne sont pas valides. Veuillez les corriger.")
            return
        
        # Sauvegarder
        if self.mapping_manager.save_mappings(product_type, mappings):
            QMessageBox.information(self, "Succès", f"Mappings {product_type} sauvegardés avec succès!")
            self.update_info()
        else:
            QMessageBox.critical(self, "Erreur", "Erreur lors de la sauvegarde des mappings.")
    
    def reset_to_defaults(self):
        """Remet les mappings aux valeurs par défaut"""
        product_type = self.current_product_type
        
        reply = QMessageBox.question(self, "Confirmation", 
                                   f"Voulez-vous vraiment remettre les mappings {product_type} aux valeurs par défaut ?",
                                   QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        
        if reply == QMessageBox.StandardButton.Yes:
            if self.mapping_manager.reset_to_defaults(product_type):
                # Recharger l'interface
                if product_type == "matelas":
                    self.matelas_mappings = self.mapping_manager.matelas_mappings.copy()
                else:
                    self.sommiers_mappings = self.mapping_manager.sommiers_mappings.copy()
                
                self.load_mappings_table()
                self.load_preview()
                self.update_info()
                
                QMessageBox.information(self, "Succès", f"Mappings {product_type} remis aux valeurs par défaut!")
            else:
                QMessageBox.critical(self, "Erreur", "Erreur lors de la réinitialisation des mappings.") 

    def change_zoom(self, delta):
        new_zoom = self.preview_widget.zoom_factor + delta
        self.preview_widget.set_zoom(new_zoom)
        self.zoom_label.setText(f"{int(self.preview_widget.zoom_factor * 100)}%") 

        zoom_btn_style = (
            "QPushButton {"
            "background: white;"
            "color: #222;"
            "border: 2px solid #222;"
            "border-radius: 8px;"
            "font-size: 20px;"
            "font-weight: bold;"
            "padding: 2px 0 2px 0;"
            "min-width: 36px;"
            "min-height: 36px;"
            "}"
            "QPushButton:hover {"
            "background: #f0f0f0;"
            "}"
        )
        self.zoom_out_btn.setStyleSheet(zoom_btn_style)
        self.zoom_in_btn.setStyleSheet(zoom_btn_style) 