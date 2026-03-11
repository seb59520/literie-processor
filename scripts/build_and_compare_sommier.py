#!/usr/bin/env python3
"""Pipeline complète : PDF -> pré-import -> Excel + comparaison."""

import argparse
import json
import logging
from dataclasses import asdict, is_dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from backend.pdf_orders_loader import PDFOrderLoader
from backend.pre_import_utils import (
    PreImportBuilder,
    convert_pdf_orders_to_sommier_configs,
)
from backend.excel_sommier_import_utils import ExcelSommierImporter

logger = logging.getLogger(__name__)


def normalize_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        return value.strip()
    return value


def compare_excels(generated: Path, reference: Path) -> Tuple[int, List[str]]:
    wb_gen = load_workbook(generated, data_only=True)
    wb_ref = load_workbook(reference, data_only=True)
    ws_gen = wb_gen.active
    ws_ref = wb_ref.active

    max_row = max(ws_gen.max_row, ws_ref.max_row)
    max_col = max(ws_gen.max_column, ws_ref.max_column)
    mismatches: List[str] = []

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v_gen = normalize_cell_value(ws_gen.cell(row=r, column=c).value)
            v_ref = normalize_cell_value(ws_ref.cell(row=r, column=c).value)
            if v_gen != v_ref:
                cell = f"{get_column_letter(c)}{r}"
                mismatches.append(f"{cell}: '{v_gen}' != '{v_ref}'")
                if len(mismatches) >= 200:
                    break
        if len(mismatches) >= 200:
            break

    wb_gen.close()
    wb_ref.close()
    return len(mismatches), mismatches


def find_order_doc(order_documents: List[Dict], key: str) -> Optional[Dict]:
    for doc in order_documents:
        doc_key = doc.get("order_number") or doc.get("file_path")
        if doc_key == key:
            return doc
    return None


def _normalize_orders(order_documents: List[Dict]) -> List[Dict]:
    normalized = []
    for doc in order_documents:
        if is_dataclass(doc):
            normalized.append(asdict(doc))
        elif isinstance(doc, dict):
            normalized.append(doc)
        else:
            # Dernier recours : s'appuyer sur __dict__
            normalized.append(dict(doc.__dict__))
    return normalized


def build_pre_import_rows(order_documents: List[Dict]) -> List[Dict]:
    normalized_docs = _normalize_orders(order_documents)
    configs = convert_pdf_orders_to_sommier_configs(normalized_docs)
    builder = PreImportBuilder()
    rows: List[Dict] = []

    for cfg in configs:
        key = cfg.get("commande_client") or cfg.get("file_path")
        order_doc = find_order_doc(normalized_docs, key) or {}
        donnees_client = {
            "nom": order_doc.get("client_normalized", "CLIENT"),
            "adresse": order_doc.get("city", ""),
        }
        mots_ops = cfg.get("mots_operation_trouves", [])
        articles = cfg.get("articles")
        pre_import = builder.build_sommier_pre_import(
            [cfg],
            donnees_client,
            mots_operation_list=mots_ops,
            articles_llm=articles,
        )
        rows.extend(pre_import)
    return rows


def main():
    parser = argparse.ArgumentParser(description="Construit et compare un fichier sommier à partir des PDF.")
    parser.add_argument("pdf_dir", help="Dossier contenant les PDF (ex: 1103)")
    parser.add_argument("semaine", help="Code semaine (ex: S11)")
    parser.add_argument("annee", help="Année utilisée dans le nom de fichier (ex: 2026)")
    parser.add_argument("reference_excel", help="Fichier Excel corrigé pour la comparaison")
    parser.add_argument(
        "--json-output",
        default="logs/pdf_orders_snapshot.json",
        help="Chemin pour sauvegarder l'extraction JSON",
    )
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="[%(levelname)s] %(message)s")

    pdf_dir = Path(args.pdf_dir)
    loader = PDFOrderLoader(pdf_dir)
    orders = loader.extract_orders()
    normalized_orders = _normalize_orders(orders)
    json_path = Path(args.json_output)
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(
        json.dumps(normalized_orders, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    logger.info("Extraction PDF sauvegardée dans %s", json_path)

    pre_import_rows = build_pre_import_rows(normalized_orders)
    logger.info("Pré-import généré (%s configurations)", len(pre_import_rows))

    importer = ExcelSommierImporter()
    created_files = importer.import_configurations(pre_import_rows, args.semaine, args.annee)
    if not created_files:
        raise RuntimeError("Aucun fichier généré")
    generated_file = Path(created_files[-1])
    logger.info("Fichier généré: %s", generated_file)

    mismatches_count, mismatches = compare_excels(generated_file, Path(args.reference_excel))
    if mismatches_count == 0:
        logger.info("Comparaison OK : aucune différence détectée")
    else:
        logger.warning("%s différences détectées. Aperçu:", mismatches_count)
        for line in mismatches[:20]:
            logger.warning("  %s", line)


if __name__ == "__main__":
    main()
