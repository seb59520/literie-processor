#!/usr/bin/env python3
"""
Script de test pour vérifier l'ouverture des fichiers Excel
"""

import sys
import os
import tempfile
from PyQt6.QtWidgets import QApplication, QMainWindow, QVBoxLayout, QPushButton, QLabel
from PyQt6.QtCore import QUrl

# Import de la méthode open_excel_file depuis app_gui
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

class TestExcelOpening(QMainWindow):
    """Fenêtre de test pour l'ouverture des fichiers Excel"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Test Ouverture Excel")
        self.setGeometry(100, 100, 400, 300)
        
        # Widget central
        central_widget = QMainWindow()
        self.setCentralWidget(central_widget)
        
        layout = QVBoxLayout()
        central_widget.setLayout(layout)
        
        # Label d'information
        info_label = QLabel("Test de l'ouverture des fichiers Excel")
        info_label.setStyleSheet("font-size: 14px; font-weight: bold; margin: 10px;")
        layout.addWidget(info_label)
        
        # Bouton pour créer un fichier Excel de test
        create_btn = QPushButton("Créer fichier Excel de test")
        create_btn.clicked.connect(self.create_test_excel)
        layout.addWidget(create_btn)
        
        # Bouton pour tester l'ouverture
        open_btn = QPushButton("Tester l'ouverture Excel")
        open_btn.clicked.connect(self.test_excel_opening)
        layout.addWidget(open_btn)
        
        # Label de statut
        self.status_label = QLabel("En attente...")
        self.status_label.setStyleSheet("margin: 10px; padding: 10px; background-color: #f0f0f0;")
        layout.addWidget(self.status_label)
        
        self.test_file_path = None
    
    def create_test_excel(self):
        """Crée un fichier Excel de test"""
        try:
            import openpyxl
            
            # Créer un workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Test"
            
            # Ajouter quelques données
            ws['A1'] = "Test"
            ws['B1'] = "Données"
            ws['A2'] = "Ligne 1"
            ws['B2'] = "Valeur 1"
            ws['A3'] = "Ligne 2"
            ws['B3'] = "Valeur 2"
            
            # Créer un fichier temporaire
            temp_dir = tempfile.gettempdir()
            self.test_file_path = os.path.join(temp_dir, "test_excel.xlsx")
            
            # Sauvegarder le fichier
            wb.save(self.test_file_path)
            
            self.status_label.setText(f"✅ Fichier Excel créé: {self.test_file_path}")
            self.status_label.setStyleSheet("margin: 10px; padding: 10px; background-color: #d4edda; color: #155724;")
            
        except ImportError:
            self.status_label.setText("❌ openpyxl non installé. Installer avec: pip install openpyxl")
            self.status_label.setStyleSheet("margin: 10px; padding: 10px; background-color: #f8d7da; color: #721c24;")
        except Exception as e:
            self.status_label.setText(f"❌ Erreur création fichier: {e}")
            self.status_label.setStyleSheet("margin: 10px; padding: 10px; background-color: #f8d7da; color: #721c24;")
    
    def test_excel_opening(self):
        """Teste l'ouverture du fichier Excel"""
        if not self.test_file_path or not os.path.exists(self.test_file_path):
            self.status_label.setText("❌ Créer d'abord un fichier Excel de test")
            self.status_label.setStyleSheet("margin: 10px; padding: 10px; background-color: #f8d7da; color: #721c24;")
            return
        
        try:
            # Importer la méthode open_excel_file depuis app_gui
            from app_gui import MatelasApp
            
            # Créer une instance temporaire pour tester la méthode
            app_instance = MatelasApp()
            
            # Créer une URL de fichier
            file_url = QUrl.fromLocalFile(self.test_file_path)
            
            # Tester la méthode open_excel_file
            app_instance.open_excel_file(file_url)
            
            self.status_label.setText("✅ Test d'ouverture Excel réussi !")
            self.status_label.setStyleSheet("margin: 10px; padding: 10px; background-color: #d4edda; color: #155724;")
            
        except Exception as e:
            self.status_label.setText(f"❌ Erreur ouverture Excel: {e}")
            self.status_label.setStyleSheet("margin: 10px; padding: 10px; background-color: #f8d7da; color: #721c24;")
            print(f"Erreur détaillée: {e}")

def main():
    """Fonction principale"""
    app = QApplication(sys.argv)
    
    # Créer la fenêtre de test
    test_window = TestExcelOpening()
    test_window.show()
    
    print("🧪 Test d'ouverture des fichiers Excel")
    print("1. Cliquez sur 'Créer fichier Excel de test'")
    print("2. Cliquez sur 'Tester l'ouverture Excel'")
    print("3. Vérifiez que le fichier s'ouvre correctement")
    
    return app.exec()

if __name__ == "__main__":
    main() 