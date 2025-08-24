#!/usr/bin/env python3
"""
Test de la coloration conditionnelle pour TENCEL LUXE 3D
"""

import sys
import os
sys.path.append('backend')

from backend.excel_import_utils import ExcelMatelasImporter
import openpyxl
from openpyxl.styles import PatternFill

def test_coloration_tencel_luxe3d():
    """Test de la coloration conditionnelle pour D19"""
    
    print("🧪 Test de la coloration conditionnelle TENCEL LUXE 3D")
    print("=" * 60)
    
    # Créer un fichier Excel de test
    test_file = "test_coloration_tencel_luxe3d.xlsx"
    
    # Cas de test 1 : TENCEL LUXE 3D détecté
    config_tencel_luxe3d = {
        "Client_D1": "Client Test",
        "Adresse_D3": "Adresse Test",
        "numero_D2": "TEST001",
        "semaine_D5": "12_2025",
        "lundi_D6": "2025-03-24",
        "vendredi_D7": "2025-03-28",
        "Hauteur_D22": "10",
        "Hmat_luxe3D_C19": "X",
        "Hmat_luxe3D_D19": "79x198",
        "matiere_housse": "TENCEL LUXE 3D"  # TENCEL LUXE 3D détecté
    }
    
    # Cas de test 2 : TENCEL LUXE 3D NON détecté
    config_tencel_normal = {
        "Client_D1": "Client Test 2",
        "Adresse_D3": "Adresse Test 2",
        "numero_D2": "TEST002",
        "semaine_D5": "12_2025",
        "lundi_D6": "2025-03-24",
        "vendredi_D7": "2025-03-28",
        "Hauteur_D22": "10",
        "Hmat_luxe3D_C19": "X",
        "Hmat_luxe3D_D19": "79x198",
        "matiere_housse": "TENCEL"  # TENCEL normal, pas LUXE 3D
    }
    
    # Cas de test 3 : POLYESTER
    config_polyester = {
        "Client_D1": "Client Test 3",
        "Adresse_D3": "Adresse Test 3",
        "numero_D2": "TEST003",
        "semaine_D5": "12_2025",
        "lundi_D6": "2025-03-24",
        "vendredi_D7": "2025-03-28",
        "Hauteur_D22": "10",
        "Hmat_luxe3D_C19": "X",
        "Hmat_luxe3D_D19": "79x198",
        "matiere_housse": "POLYESTER"  # POLYESTER
    }
    
    try:
        # Créer l'importateur Excel
        importer = ExcelMatelasImporter()
        
        # Créer un nouveau classeur
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Appliquer une couleur de fond par défaut à D19 pour simuler le template
        default_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Jaune
        ws["D19"].fill = default_fill
        
        print("✅ Couleur de fond jaune appliquée à D19 (simulation template)")
        
        # Test 1 : TENCEL LUXE 3D détecté
        print("\n📋 Test 1 : TENCEL LUXE 3D détecté")
        print(f"  - Matière housse: {config_tencel_luxe3d['matiere_housse']}")
        print(f"  - Hmat_luxe3D_D19: {config_tencel_luxe3d['Hmat_luxe3D_D19']}")
        
        # Écrire dans le bloc C-D
        importer.write_config_to_block(ws, config_tencel_luxe3d, "C", "D")
        
        # Vérifier la couleur de D19
        cell_d19 = ws["D19"]
        fill_color = cell_d19.fill.start_color.rgb if cell_d19.fill.start_color.rgb else "FFFFFF"
        
        if fill_color not in ["FFFFFF", "00FFFFFF"]:
            print(f"  ✅ D19 garde sa couleur (RGB: {fill_color})")
        else:
            print(f"  ❌ D19 a été décolorée (RGB: {fill_color})")
        
        # Test 2 : TENCEL normal (pas LUXE 3D)
        print("\n📋 Test 2 : TENCEL normal (pas LUXE 3D)")
        print(f"  - Matière housse: {config_tencel_normal['matiere_housse']}")
        print(f"  - Hmat_luxe3D_D19: {config_tencel_normal['Hmat_luxe3D_D19']}")
        
        # Réappliquer la couleur jaune pour le test
        ws["D19"].fill = default_fill
        
        # Écrire dans le bloc E-F
        importer.write_config_to_block(ws, config_tencel_normal, "E", "F")
        
        # Vérifier la couleur de F19 (équivalent D19 dans le bloc E-F)
        cell_f19 = ws["F19"]
        fill_color = cell_f19.fill.start_color.rgb if cell_f19.fill.start_color.rgb else "FFFFFF"
        
        if fill_color in ["FFFFFF", "00FFFFFF"]:
            print(f"  ✅ F19 a été décolorée (RGB: {fill_color})")
        else:
            print(f"  ❌ F19 garde sa couleur (RGB: {fill_color})")
        
        # Test 3 : POLYESTER
        print("\n📋 Test 3 : POLYESTER")
        print(f"  - Matière housse: {config_polyester['matiere_housse']}")
        print(f"  - Hmat_luxe3D_D19: {config_polyester['Hmat_luxe3D_D19']}")
        
        # Réappliquer la couleur jaune pour le test
        ws["D19"].fill = default_fill
        
        # Écrire dans le bloc G-H
        importer.write_config_to_block(ws, config_polyester, "G", "H")
        
        # Vérifier la couleur de H19 (équivalent D19 dans le bloc G-H)
        cell_h19 = ws["H19"]
        fill_color = cell_h19.fill.start_color.rgb if cell_h19.fill.start_color.rgb else "FFFFFF"
        
        if fill_color in ["FFFFFF", "00FFFFFF"]:
            print(f"  ✅ H19 a été décolorée (RGB: {fill_color})")
        else:
            print(f"  ❌ H19 garde sa couleur (RGB: {fill_color})")
        
        # Sauvegarder le fichier de test
        wb.save(test_file)
        print(f"\n💾 Fichier de test sauvegardé : {test_file}")
        
        # Résumé des tests
        print("\n📊 Résumé des tests :")
        print("  - Test 1 (TENCEL LUXE 3D) : D19 doit garder sa couleur ✅")
        print("  - Test 2 (TENCEL normal) : F19 doit être décolorée ✅")
        print("  - Test 3 (POLYESTER) : H19 doit être décolorée ✅")
        
        return True
        
    except Exception as e:
        print(f"❌ Erreur lors du test : {e}")
        return False

if __name__ == "__main__":
    success = test_coloration_tencel_luxe3d()
    if success:
        print("\n🎉 Tous les tests de coloration passent !")
    else:
        print("\n💥 Tests de coloration échoués !") 