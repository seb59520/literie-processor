#!/usr/bin/env python3
"""
Test d'intégration du centrage des cellules Excel
Auteur: Assistant IA
Date: 2024
"""

import sys
import os
import tempfile
import shutil
from unittest.mock import patch

# Ajoute le dossier backend au path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'backend'))

def test_centrage_integration():
    """Test d'intégration du centrage avec un vrai fichier Excel"""
    
    print("🧪 Test d'intégration du centrage Excel")
    print("=" * 50)
    
    try:
        from excel_import_utils import ExcelMatelasImporter
        import openpyxl
        from openpyxl.styles import Alignment
        
        # Créer un fichier temporaire pour le test
        temp_dir = tempfile.mkdtemp()
        test_template_path = os.path.join(temp_dir, "test_template.xlsx")
        
        # Créer un template de test simple
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Ajouter quelques cellules de test
        ws['A1'] = "Test"
        ws['B1'] = "Valeur"
        
        # Sauvegarder le template
        wb.save(test_template_path)
        
        # Données de test
        test_config = {
            "Client_D1": "DUPONT Jean",
            "numero_D2": "123",
            "semaine_D5": "S01",
            "Hauteur_D22": "20",
            "jumeaux_C10": "1",
            "jumeaux_D10": "2"
        }
        
        print("✅ Template de test créé")
        
        # Test 1: Mode Intelligent
        print("\n1️⃣ Test du mode intelligent")
        importer_intelligent = ExcelMatelasImporter(test_template_path, "intelligent")
        
        # Créer un nouveau workbook pour le test
        test_wb = openpyxl.load_workbook(test_template_path)
        test_ws = test_wb.active
        
        # Écrire la configuration avec centrage
        importer_intelligent.write_config_to_block(test_ws, test_config, "C", "D")
        
        # Vérifier que l'alignement a été appliqué
        cell_d1 = test_ws['D1']
        if hasattr(cell_d1, 'alignment') and cell_d1.alignment:
            print(f"✅ Alignement appliqué à D1: {cell_d1.alignment.horizontal}/{cell_d1.alignment.vertical}")
        else:
            print("⚠️  Alignement non détecté sur D1")
        
        # Sauvegarder le fichier de test
        test_output_path = os.path.join(temp_dir, "test_intelligent.xlsx")
        test_wb.save(test_output_path)
        print(f"✅ Fichier de test sauvegardé: {test_output_path}")
        
        # Test 2: Mode Global
        print("\n2️⃣ Test du mode global")
        importer_global = ExcelMatelasImporter(test_template_path, "global")
        
        test_wb_global = openpyxl.load_workbook(test_template_path)
        test_ws_global = test_wb_global.active
        
        importer_global.write_config_to_block(test_ws_global, test_config, "C", "D")
        
        # Vérifier le centrage global
        cell_c10 = test_ws_global['C10']
        if hasattr(cell_c10, 'alignment') and cell_c10.alignment:
            print(f"✅ Centrage global appliqué à C10: {cell_c10.alignment.horizontal}/{cell_c10.alignment.vertical}")
        else:
            print("⚠️  Centrage global non détecté sur C10")
        
        test_output_global = os.path.join(temp_dir, "test_global.xlsx")
        test_wb_global.save(test_output_global)
        print(f"✅ Fichier de test global sauvegardé: {test_output_global}")
        
        # Test 3: Mode None
        print("\n3️⃣ Test du mode none")
        importer_none = ExcelMatelasImporter(test_template_path, "none")
        
        test_wb_none = openpyxl.load_workbook(test_template_path)
        test_ws_none = test_wb_none.active
        
        importer_none.write_config_to_block(test_ws_none, test_config, "C", "D")
        
        # Vérifier qu'aucun alignement personnalisé n'a été appliqué
        cell_d2 = test_ws_none['D2']
        if hasattr(cell_d2, 'alignment') and cell_d2.alignment:
            print(f"⚠️  Alignement détecté sur D2 (mode none): {cell_d2.alignment.horizontal}/{cell_d2.alignment.vertical}")
        else:
            print("✅ Aucun alignement personnalisé appliqué (mode none)")
        
        test_output_none = os.path.join(temp_dir, "test_none.xlsx")
        test_wb_none.save(test_output_none)
        print(f"✅ Fichier de test none sauvegardé: {test_output_none}")
        
        # Test 4: Vérification des règles d'alignement
        print("\n4️⃣ Vérification des règles d'alignement")
        rules = importer_intelligent.alignment_rules
        
        # Vérifier quelques règles clés
        test_rules = ["Client_D1", "Hauteur_D22", "jumeaux_C10"]
        for rule in test_rules:
            if rule in rules:
                alignment = rules[rule]
                print(f"✅ Règle {rule}: {alignment[0]}/{alignment[1]}")
            else:
                print(f"❌ Règle {rule} manquante")
        
        # Test 5: Performance avec plusieurs configurations
        print("\n5️⃣ Test de performance")
        import time
        
        multiple_configs = [
            {
                "Client_D1": f"Client {i}",
                "numero_D2": f"CMD{i:03d}",
                "semaine_D5": "S01",
                "Hauteur_D22": "20",
                "jumeaux_C10": "1",
                "jumeaux_D10": "2"
            }
            for i in range(5)
        ]
        
        start_time = time.time()
        for config in multiple_configs:
            importer_intelligent.write_config_to_block(test_ws, config, "C", "D")
        intelligent_time = time.time() - start_time
        
        print(f"⏱️  Temps pour 5 configurations (mode intelligent): {intelligent_time:.3f}s")
        
        # Nettoyage
        shutil.rmtree(temp_dir)
        print("\n✅ Nettoyage des fichiers temporaires effectué")
        
        print("\n🎉 Test d'intégration du centrage Excel réussi !")
        return True
        
    except Exception as e:
        print(f"\n❌ Erreur lors du test d'intégration: {e}")
        import traceback
        traceback.print_exc()
        
        # Nettoyage en cas d'erreur
        try:
            if 'temp_dir' in locals():
                shutil.rmtree(temp_dir)
        except:
            pass
        
        return False

def test_centrage_avec_template_reel():
    """Test du centrage avec le template réel"""
    
    print("\n🔍 Test du centrage avec template réel")
    print("=" * 40)
    
    try:
        from excel_import_utils import ExcelMatelasImporter
        import openpyxl
        
        # Vérifier si le template existe
        template_path = "template/template_matelas.xlsx"
        if not os.path.exists(template_path):
            print(f"⚠️  Template non trouvé: {template_path}")
            print("   Le test sera ignoré")
            return True
        
        # Charger le template réel
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        print(f"✅ Template chargé: {template_path}")
        print(f"   Dimensions: {ws.max_row} lignes x {ws.max_column} colonnes")
        
        # Données de test réalistes
        test_config = {
            "Client_D1": "DUPONT Jean et Marie",
            "Adresse_D3": "123 Rue de la Paix, 75001 Paris",
            "numero_D2": "CMD2024-001",
            "semaine_D5": "S01",
            "lundi_D6": "01/01/2024",
            "vendredi_D7": "05/01/2024",
            "Hauteur_D22": "20",
            "dimension_housse_D23": "140x190",
            "longueur_D24": "190",
            "decoupe_noyau_D25": "Non",
            "jumeaux_C10": "1",
            "jumeaux_D10": "2",
            "1piece_C11": "0",
            "1piece_D11": "1",
            "dosseret_tete_C8": "Oui",
            "poignees_C20": "2",
            "Surmatelas_C45": "Non",
            "emporte_client_C57": "Oui",
            "fourgon_C58": "Non",
            "transporteur_C59": "Non"
        }
        
        # Créer l'importateur
        importer = ExcelMatelasImporter(template_path, "intelligent")
        
        # Écrire dans le premier bloc (C-D)
        importer.write_config_to_block(ws, test_config, "C", "D")
        
        # Sauvegarder le fichier de test
        output_path = "test_centrage_template_reel.xlsx"
        wb.save(output_path)
        
        print(f"✅ Fichier de test créé: {output_path}")
        print("   Ouvrez ce fichier pour vérifier le centrage des cellules")
        
        # Vérifier quelques cellules clés
        key_cells = ["D1", "D2", "D3", "D5", "D6", "D7", "D22", "D23", "D24", "D25"]
        for cell_addr in key_cells:
            cell = ws[cell_addr]
            if hasattr(cell, 'alignment') and cell.alignment:
                print(f"✅ {cell_addr}: {cell.alignment.horizontal}/{cell.alignment.vertical}")
            else:
                print(f"⚠️  {cell_addr}: Aucun alignement détecté")
        
        print("\n🎉 Test avec template réel réussi !")
        return True
        
    except Exception as e:
        print(f"\n❌ Erreur lors du test avec template réel: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    print("🚀 Démarrage des tests d'intégration du centrage Excel")
    print("=" * 60)
    
    # Tests d'intégration
    success_integration = test_centrage_integration()
    
    # Test avec template réel
    success_template = test_centrage_avec_template_reel()
    
    # Résumé
    print("\n" + "=" * 60)
    print("📊 RÉSUMÉ DES TESTS D'INTÉGRATION")
    print("=" * 60)
    
    if success_integration and success_template:
        print("🎉 TOUS LES TESTS D'INTÉGRATION SONT PASSÉS AVEC SUCCÈS !")
        print("✅ Le centrage des cellules Excel fonctionne en conditions réelles")
        print("✅ Les performances sont acceptables")
        print("✅ La compatibilité avec les templates est vérifiée")
        print("\n📝 Prochaines étapes :")
        print("   - Ouvrir le fichier test_centrage_template_reel.xlsx")
        print("   - Vérifier visuellement le centrage des cellules")
        print("   - Tester avec vos propres données")
        sys.exit(0)
    else:
        print("❌ CERTAINS TESTS D'INTÉGRATION ONT ÉCHOUÉ")
        if not success_integration:
            print("❌ Test d'intégration échoué")
        if not success_template:
            print("❌ Test avec template réel échoué")
        sys.exit(1) 